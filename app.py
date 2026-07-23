# app.py
# ================================================================================
#                            H Y D R O   D O M U S   P Y
# ================================================================================
#                    Análisis Hidráulico para Redes Internas
#                          Interfaz Web con Streamlit
# ================================================================================
#                               Versión 1.0.0 - Web
# ================================================================================

import streamlit as st
import ezdxf
import numpy as np
import networkx as nx
from scipy.optimize import fsolve
import plotly.graph_objects as go
from plotly.subplots import make_subplots
import tempfile
import os
import json
import base64
import io
from datetime import datetime
from collections import deque
from dataclasses import dataclass
from typing import List, Tuple, Dict, Optional
import math
import pandas as pd

# ================================================================================
# CONFIGURACIÓN DE PÁGINA
# ================================================================================
st.set_page_config(
    page_title="Hydro Domus Py - Análisis Hidráulico",
    page_icon="💧",
    layout="wide",
    initial_sidebar_state="expanded"
)

# ================================================================================
# CONFIGURACIÓN GLOBAL
# ================================================================================
UNIDADES_GASTO = {
    "LAVAMANOS": {"ug": 1, "caudal_unitario": 0.1, "color": "#3498db", "icono": "🚰"},
    "LAVADERO": {"ug": 3, "caudal_unitario": 0.3, "color": "#2ecc71", "icono": "🧺"},
    "SANITARIO": {"ug": 3, "caudal_unitario": 0.3, "color": "#e67e22", "icono": "🚽"},
    "DUCHA": {"ug": 2, "caudal_unitario": 0.2, "color": "#1abc9c", "icono": "🚿"},
    "LAVADORA": {"ug": 3, "caudal_unitario": 0.3, "color": "#9b59b6", "icono": "👕"},
    "TANQUE": {"ug": 5, "caudal_unitario": 0.5, "color": "#f39c12", "icono": "🗄️"},
    "LAVAPLATOS": {"ug": 2, "caudal_unitario": 0.2, "color": "#e74c3c", "icono": "🍽️"},
    "GRIFO_JARDIN": {"ug": 2, "caudal_unitario": 0.2, "color": "#27ae60", "icono": "🌿"},
}

DIAMETROS_PAVCO = {
    "1/2": 18.18,
    "3/4": 23.63,
    "1": 30.20,
    "1_1/2": 43.68,
    "2": 55.68,
    "3": 82.04,
    "4": 103.42,
}

EPSILON_PVC_MM = 0.0015
GRAVEDAD = 9.81
VISCOSIDAD_AGUA = 1e-6
PRESION_MIN_NORMA = 5.0
TOLERANCIA_CONEXION_M = 0.01

LEQ_ACCESORIOS = {
    'Tee': 1.5,
    'Codo_90': 0.8,
    'Codo_45': 0.5,
    'Reduccion': 0.5,
    'Valvula_Compuerta': 0.3,
    'Valvula_Globo': 1.5,
    'Valvula_Check': 2.5,
    'Valvula_Esfera': 0.2,
}

TIPOS_OCUPACION_AGUA = {
    "Vivienda Unifamiliar": {
        "descripcion": "Casas, residencias unifamiliares",
        "formula": "Q = 0.2 * √UG + 0.5",
        "coeficiente_a": 0.20,
        "coeficiente_b": 0.50,
        "color": "#3498db"
    },
    "Vivienda Multifamiliar": {
        "descripcion": "Edificios de apartamentos (hasta 20 unidades)",
        "formula": "Q = 0.25 * √UG + 0.6",
        "coeficiente_a": 0.25,
        "coeficiente_b": 0.60,
        "color": "#2ecc71"
    },
    "Edificio de Oficinas": {
        "descripcion": "Oficinas, edificios administrativos",
        "formula": "Q = 0.15 * √UG + 0.3",
        "coeficiente_a": 0.15,
        "coeficiente_b": 0.30,
        "color": "#e67e22"
    },
    "Hotel / Hostería": {
        "descripcion": "Hoteles, hostales, alojamientos",
        "formula": "Q = 0.18 * √UG + 0.4",
        "coeficiente_a": 0.18,
        "coeficiente_b": 0.40,
        "color": "#f39c12"
    },
    "Hospital / Clínica": {
        "descripcion": "Centros médicos, hospitales",
        "formula": "Q = 0.22 * √UG + 0.7",
        "coeficiente_a": 0.22,
        "coeficiente_b": 0.70,
        "color": "#e74c3c"
    },
    "Centro Comercial": {
        "descripcion": "Centros comerciales, tiendas",
        "formula": "Q = 0.12 * √UG + 0.4",
        "coeficiente_a": 0.12,
        "coeficiente_b": 0.40,
        "color": "#9b59b6"
    },
    "Colegio / Universidad": {
        "descripcion": "Instituciones educativas",
        "formula": "Q = 0.13 * √UG + 0.35",
        "coeficiente_a": 0.13,
        "coeficiente_b": 0.35,
        "color": "#1abc9c"
    },
    "Restaurante / Cafetería": {
        "descripcion": "Establecimientos de comida",
        "formula": "Q = 0.16 * √UG + 0.45",
        "coeficiente_a": 0.16,
        "coeficiente_b": 0.45,
        "color": "#e84393"
    },
    "Gimnasio / Deportivo": {
        "descripcion": "Centros deportivos, gimnasios",
        "formula": "Q = 0.20 * √UG + 0.55",
        "coeficiente_a": 0.20,
        "coeficiente_b": 0.55,
        "color": "#00cec9"
    },
    "Ancianato / Geriátrico": {
        "descripcion": "Residencias de ancianos",
        "formula": "Q = 0.17 * √UG + 0.4",
        "coeficiente_a": 0.17,
        "coeficiente_b": 0.40,
        "color": "#fd79a8"
    }
}

# ================================================================================
# UNIDADES DE DIBUJO
# ================================================================================
UNIDADES_DIBUJO = {
    "mm": {"nombre": "Milímetros", "factor": 1000, "icono": "📏"},
    "cm": {"nombre": "Centímetros", "factor": 100, "icono": "📐"},
    "m": {"nombre": "Metros", "factor": 1, "icono": "📏"},
    "in": {"nombre": "Pulgadas", "factor": 0.0254, "icono": "📐"},
    "ft": {"nombre": "Pies", "factor": 0.3048, "icono": "📏"},
}

# ================================================================================
# MODELOS DE DATOS
# ================================================================================
@dataclass
class Nodo:
    id: int
    x: float
    y: float
    z: float
    es_entrada: bool = False
    tipo_aparato: str = ""
    valvula_tipo: str = ""
    valvula_cerrada: bool = False
    demanda_lps: float = 0.0
    presion_mca: Optional[float] = None

@dataclass
class Accesorio:
    id: int
    tipo: str
    nodo_id: int
    longitud_equivalente_m: float = 0.0
    perdida_mca: float = 0.0
    
    def calcular_perdida(self, f_friccion: float, diametro_m: float, velocidad_ms: float) -> float:
        if self.longitud_equivalente_m > 0 and velocidad_ms > 0:
            self.perdida_mca = f_friccion * (self.longitud_equivalente_m / diametro_m) * (velocidad_ms**2 / (2 * GRAVEDAD))
        return self.perdida_mca

@dataclass
class Tuberia:
    id: int
    nodo_inicio: int
    nodo_fin: int
    longitud_m: float
    diametro_mm: float = 21.0
    caudal_lps: float = 0.0
    velocidad_ms: float = 0.0
    perdida_mca: float = 0.0
    f_friccion: float = 0.02
    
    @property
    def diametro_m(self): return self.diametro_mm / 1000.0
    @property
    def diametro_pulg(self): return self.diametro_mm / 25.4
    @property
    def diametro_nominal_pulg(self):
        mapeo = {18.18: "1/2", 23.63: "3/4", 30.20: "1", 43.68: "1-1/2", 55.68: "2", 82.04: "3", 103.42: "4"}
        if self.diametro_mm in mapeo:
            return mapeo[self.diametro_mm]
        cercano = min(mapeo.keys(), key=lambda x: abs(x - self.diametro_mm))
        return mapeo[cercano]
    @property
    def area_m2(self): return np.pi * (self.diametro_m / 2) ** 2
    
    def calcular_velocidad(self):
        if self.caudal_lps > 0:
            self.velocidad_ms = (self.caudal_lps / 1000.0) / self.area_m2
        return self.velocidad_ms
    
    def colebrook_white(self, Re):
        epsilon = EPSILON_PVC_MM / 1000.0
        D = self.diametro_m
        def ecuacion(f):
            return 1/np.sqrt(f) + 2 * np.log10(epsilon/(3.71*D) + 2.51/(Re*np.sqrt(f)))
        try:
            f_solution = fsolve(ecuacion, 0.02)[0]
            return max(0.008, min(0.05, f_solution))
        except:
            return 0.02
    
    def calcular_perdida(self, accesorios_por_nodo: Dict[int, List[Accesorio]] = None) -> float:
        if self.caudal_lps == 0:
            self.perdida_mca = 0.0
            return 0.0
        
        V = self.velocidad_ms
        D = self.diametro_m
        Re = V * D / VISCOSIDAD_AGUA
        f = self.colebrook_white(Re)
        self.f_friccion = f
        
        hf_friccion = f * (self.longitud_m / D) * (V**2 / (2 * GRAVEDAD))
        hf_accesorios = 0.0
        if accesorios_por_nodo:
            for acc in accesorios_por_nodo.get(self.nodo_inicio, []):
                hf_accesorios += acc.calcular_perdida(f, D, V)
            for acc in accesorios_por_nodo.get(self.nodo_fin, []):
                hf_accesorios += acc.calcular_perdida(f, D, V)
        
        self.perdida_mca = hf_friccion + hf_accesorios
        return self.perdida_mca

# ================================================================================
# FUNCIONES AUXILIARES
# ================================================================================
def caudal_por_ug(ug_total: float, tipo_ocupacion: str = None) -> float:
    if ug_total <= 0:
        return 0.0
    
    if tipo_ocupacion is None:
        tipo_ocupacion = st.session_state.get('tipo_ocupacion', "Vivienda Unifamiliar")
    
    config = TIPOS_OCUPACION_AGUA.get(tipo_ocupacion, TIPOS_OCUPACION_AGUA["Vivienda Unifamiliar"])
    a = config.get("coeficiente_a", 0.2)
    b = config.get("coeficiente_b", 0.5)
    return a * (ug_total ** 0.5) + b

def diametro_a_numero(diam):
    mapeo = {'1/2': 0.5, '3/4': 0.75, '1': 1.0, '1-1/2': 1.5, '2': 2.0, '3': 3.0, '4': 4.0}
    return mapeo.get(diam, 999)

# ================================================================================
# CLASE RED HIDRÁULICA
# ================================================================================
class RedHidraulica:
    def __init__(self):
        self.nodos: Dict[int, Nodo] = {}
        self.tuberias: Dict[int, Tuberia] = {}
        self.accesorios: List[Accesorio] = []
        self.grafo = nx.Graph()
        self.nodo_entrada_id = None
    
    def agregar_nodo(self, nodo):
        self.nodos[nodo.id] = nodo
        self.grafo.add_node(nodo.id)
    
    def agregar_tuberia(self, tubo):
        self.tuberias[tubo.id] = tubo
        self.grafo.add_edge(tubo.nodo_inicio, tubo.nodo_fin)
    
    def _find_tuberia(self, n1, n2):
        for tid, t in self.tuberias.items():
            if (t.nodo_inicio == n1 and t.nodo_fin == n2) or (t.nodo_inicio == n2 and t.nodo_fin == n1):
                return tid
        return None

    def calcular_ug_acumulada(self, alcanzables=None):
        if self.nodo_entrada_id is None:
            return {}
        
        if alcanzables is None:
            alcanzables = set(self.nodos.keys())
        
        ug_por_nodo = {}
        for nid, nodo in self.nodos.items():
            ug = 0
            if nid in alcanzables and nodo.tipo_aparato and nodo.tipo_aparato in UNIDADES_GASTO:
                ug = UNIDADES_GASTO[nodo.tipo_aparato]["ug"]
            ug_por_nodo[nid] = ug
        
        padres = {self.nodo_entrada_id: None}
        orden_bfs = [self.nodo_entrada_id]
        cola = deque([self.nodo_entrada_id])
        
        while cola:
            nodo = cola.popleft()
            if nodo not in alcanzables:
                continue
            for vecino in self.grafo.neighbors(nodo):
                if vecino not in padres and vecino in alcanzables:
                    padres[vecino] = nodo
                    orden_bfs.append(vecino)
                    cola.append(vecino)
        
        ug_acumulada = ug_por_nodo.copy()
        for nodo in reversed(orden_bfs):
            if nodo == self.nodo_entrada_id:
                continue
            padre = padres.get(nodo)
            if padre is not None:
                ug_acumulada[padre] = ug_acumulada.get(padre, 0) + ug_acumulada[nodo]
        
        return ug_acumulada
    
    def calcular_angulo_entre_tuberias(self, tuberia1, tuberia2, nodo_comun):
        nodo = self.nodos[nodo_comun]
        otro1 = tuberia1.nodo_fin if tuberia1.nodo_inicio == nodo_comun else tuberia1.nodo_inicio
        otro2 = tuberia2.nodo_fin if tuberia2.nodo_inicio == nodo_comun else tuberia2.nodo_inicio
        n1, n2 = self.nodos[otro1], self.nodos[otro2]
        v1 = np.array([n1.x - nodo.x, n1.y - nodo.y, n1.z - nodo.z])
        v2 = np.array([n2.x - nodo.x, n2.y - nodo.y, n2.z - nodo.z])
        norm1, norm2 = np.linalg.norm(v1), np.linalg.norm(v2)
        if norm1 == 0 or norm2 == 0:
            return 0.0
        v1, v2 = v1 / norm1, v2 / norm2
        angulo_deg = np.degrees(np.arccos(np.clip(np.dot(v1, v2), -1.0, 1.0)))
        return angulo_deg if angulo_deg <= 90 else 180 - angulo_deg
        
    def detectar_accesorios(self):
        self.accesorios = []
        next_id = 0
        
        for nid in self.nodos.keys():
            tuberias_conectadas = [t for t in self.tuberias.values() 
                                   if t.nodo_inicio == nid or t.nodo_fin == nid]
            grado = len(tuberias_conectadas)
            
            if grado >= 3:
                self.accesorios.append(Accesorio(id=next_id, tipo='Tee', nodo_id=nid,
                                                  longitud_equivalente_m=LEQ_ACCESORIOS['Tee']))
                next_id += 1
            
            if grado >= 2:
                for i in range(len(tuberias_conectadas)):
                    for j in range(i + 1, len(tuberias_conectadas)):
                        if tuberias_conectadas[i].diametro_mm != tuberias_conectadas[j].diametro_mm:
                            self.accesorios.append(Accesorio(id=next_id, tipo='Reduccion', nodo_id=nid,
                                                              longitud_equivalente_m=LEQ_ACCESORIOS['Reduccion']))
                            next_id += 1
                            break
                    else:
                        continue
                    break
            
            if grado == 2:
                d1, d2 = tuberias_conectadas[0].diametro_mm, tuberias_conectadas[1].diametro_mm
                if d1 == d2:
                    angulo = self.calcular_angulo_entre_tuberias(tuberias_conectadas[0], tuberias_conectadas[1], nid)
                    
                    if 35 <= angulo <= 55:
                        tipo_codo, leq = 'Codo_45', LEQ_ACCESORIOS['Codo_45']
                    elif 70 <= angulo <= 110:
                        tipo_codo, leq = 'Codo_90', LEQ_ACCESORIOS['Codo_90']
                    else:
                        continue
                    
                    self.accesorios.append(Accesorio(id=next_id, tipo=tipo_codo, nodo_id=nid, longitud_equivalente_m=leq))
                    next_id += 1
        
        for nid, nodo in self.nodos.items():
            if nodo.valvula_tipo and not nodo.valvula_cerrada:
                leq_valvula = LEQ_ACCESORIOS.get(f'Valvula_{nodo.valvula_tipo}', 0.5)
                if not any(a.nodo_id == nid and 'Valvula' in a.tipo for a in self.accesorios):
                    self.accesorios.append(Accesorio(id=next_id, tipo=f'Valvula_{nodo.valvula_tipo}',
                                                      nodo_id=nid, longitud_equivalente_m=leq_valvula))
                    next_id += 1
        
        return self.accesorios

# ================================================================================
# CLASE ANALIZADOR HIDRÁULICO
# ================================================================================
class HydraulicAnalyzer:
    def __init__(self, red, callback_progreso=None, diametro_maximo=None):
        self.red = red
        self.diam_comerciales_original = sorted(DIAMETROS_PAVCO.values())
        self.diametro_maximo = diametro_maximo
        self.callback = callback_progreso
        
        self.vel_min = st.session_state.get('vel_min', 0.5)
        self.vel_max = st.session_state.get('vel_max', 2.0)
        self.presion_entrada = st.session_state.get('presion_entrada', 15.0)
        
        if self.diametro_maximo is not None:
            self.diam_comerciales = [d for d in self.diam_comerciales_original if d <= self.diametro_maximo]
        else:
            self.diam_comerciales = self.diam_comerciales_original.copy()
        
        if not self.diam_comerciales:
            self.diam_comerciales = [min(self.diam_comerciales_original)]
        
        diam_max = max(self.diam_comerciales)
        for t in self.red.tuberias.values():
            if t.diametro_mm > diam_max:
                t.diametro_mm = diam_max
    
    def asignar_caudales(self):
        if self.red.nodo_entrada_id is None:
            return
        
        alcanzables = set()
        cola = deque([self.red.nodo_entrada_id])
        
        while cola:
            nodo_actual = cola.popleft()
            if nodo_actual in alcanzables:
                continue
            alcanzables.add(nodo_actual)
            
            nodo_obj = self.red.nodos.get(nodo_actual)
            if nodo_obj and nodo_obj.valvula_cerrada:
                continue
            
            for vecino in self.red.grafo.neighbors(nodo_actual):
                if vecino not in alcanzables:
                    cola.append(vecino)
        
        ug_acumulada = self.red.calcular_ug_acumulada(alcanzables=alcanzables)
        tipo_ocupacion = st.session_state.get('tipo_ocupacion', "Vivienda Unifamiliar")
        
        for t in self.red.tuberias.values():
            try:
                dist_inicio = nx.shortest_path_length(self.red.grafo, self.red.nodo_entrada_id, t.nodo_inicio)
                dist_fin = nx.shortest_path_length(self.red.grafo, self.red.nodo_entrada_id, t.nodo_fin)
            except:
                dist_inicio, dist_fin = 0, 1
            
            nodo_aguas_abajo = t.nodo_fin if dist_fin > dist_inicio else t.nodo_inicio
            
            if nodo_aguas_abajo not in alcanzables:
                t.caudal_lps = 0
            else:
                t.caudal_lps = caudal_por_ug(ug_acumulada.get(nodo_aguas_abajo, 0), tipo_ocupacion)
    
    def optimizar_diametros(self):
        diam_min, diam_max = min(self.diam_comerciales), max(self.diam_comerciales)
        
        for t in self.red.tuberias.values():
            if t.diametro_mm > diam_max:
                t.diametro_mm = diam_max
            
            caudal = t.caudal_lps
            if caudal > 4.0:
                diam_min_sugerido = 55.68
            elif caudal > 2.5:
                diam_min_sugerido = 43.68
            elif caudal > 1.5:
                diam_min_sugerido = 30.20
            elif caudal > 0.8:
                diam_min_sugerido = 23.63
            else:
                diam_min_sugerido = 18.18
            
            cercano = min(self.diam_comerciales, key=lambda x: abs(x - diam_min_sugerido))
            diam_min_sugerido = min(cercano, diam_max)
            
            if t.diametro_mm < diam_min_sugerido:
                t.diametro_mm = diam_min_sugerido
        
        for iteracion in range(15):
            cambios = False
            for t in self.red.tuberias.values():
                v = t.calcular_velocidad()
                d_actual = t.diametro_mm
                
                if v < self.vel_min and d_actual > diam_min:
                    idx = self.diam_comerciales.index(d_actual)
                    nuevo_diam = self.diam_comerciales[max(0, idx-1)]
                    if nuevo_diam != d_actual:
                        t.diametro_mm = nuevo_diam
                        cambios = True
                elif v > self.vel_max and d_actual < diam_max:
                    idx = self.diam_comerciales.index(d_actual)
                    nuevo_diam = self.diam_comerciales[min(len(self.diam_comerciales)-1, idx+1)]
                    if nuevo_diam != d_actual:
                        t.diametro_mm = nuevo_diam
                        cambios = True
            
            if not cambios:
                break
    
    def calcular_presiones(self):
        if self.red.nodo_entrada_id is None:
            return
        
        entrada_id = self.red.nodo_entrada_id
        
        alcanzables = set()
        cola = deque([entrada_id])
        
        while cola:
            nodo_actual = cola.popleft()
            if nodo_actual in alcanzables:
                continue
            alcanzables.add(nodo_actual)
            
            nodo_obj = self.red.nodos.get(nodo_actual)
            if nodo_obj and nodo_obj.valvula_cerrada:
                continue
            
            for vecino in self.red.grafo.neighbors(nodo_actual):
                if vecino not in alcanzables:
                    cola.append(vecino)
        
        self.red.detectar_accesorios()
        
        accesorios_por_nodo = {}
        for acc in self.red.accesorios:
            accesorios_por_nodo.setdefault(acc.nodo_id, []).append(acc)
        
        for t in self.red.tuberias.values():
            t.calcular_velocidad()
            t.calcular_perdida(accesorios_por_nodo)
        
        for nodo in self.red.nodos.values():
            nodo.presion_mca = None
        
        self.red.nodos[entrada_id].presion_mca = self.presion_entrada
        
        queue = deque([entrada_id])
        visitados = set()
        
        while queue:
            nodo_actual = queue.popleft()
            if nodo_actual in visitados:
                continue
            
            visitados.add(nodo_actual)
            p_act = self.red.nodos[nodo_actual].presion_mca
            
            if p_act is None:
                continue
            
            nodo_obj = self.red.nodos.get(nodo_actual)
            if nodo_obj and nodo_obj.valvula_cerrada:
                continue
            
            for vecino in self.red.grafo.neighbors(nodo_actual):
                if vecino in visitados:
                    continue
                
                tid = self.red._find_tuberia(nodo_actual, vecino)
                if tid is None:
                    continue
                
                t = self.red.tuberias[tid]
                hf = t.calcular_perdida(accesorios_por_nodo)
                dz = self.red.nodos[vecino].z - self.red.nodos[nodo_actual].z
                p_vecino = p_act - hf - dz
                
                self.red.nodos[vecino].presion_mca = p_vecino
                queue.append(vecino)
    
    def ejecutar(self):
        self.asignar_caudales()
        self.optimizar_diametros()
        self.red.detectar_accesorios()
        self.calcular_presiones()

# ================================================================================
# CLASE LECTOR DXF
# ================================================================================

class DXFReader:
    def __init__(self, archivo):
        """Inicializa el lector DXF"""
        self.doc = ezdxf.readfile(archivo)
        self.msp = self.doc.modelspace()
        self.log_callback = None
    
    def set_log_callback(self, callback):
        """Establece una función callback para logging"""
        self.log_callback = callback
    
    def log_message(self, message, level="INFO"):
        """Envía mensaje de log al callback si existe"""
        if self.log_callback:
            self.log_callback(message, level)
        else:
            print(f"[{level}] {message}")
    
    def obtener_layers(self):
        """Obtiene la lista de layers del DXF"""
        return [layer.dxf.name for layer in self.doc.layers]
    
    def extraer_lineas(self, layers):
        """
        Extrae todas las líneas y polilíneas de los layers seleccionados.
        CORREGIDO: Manejo correcto de points() con context manager
        Soporta: LINE, LWPOLYLINE, POLYLINE, 3DPOLYLINE
        """
        lineas = []
        layers_set = set(layers)
        
        stats = {
            'LINE': 0,
            'LWPOLYLINE': 0,
            'POLYLINE': 0,
            '3DPOLYLINE': 0,
            'total_segmentos': 0
        }
        
        for entity in self.msp:
            dxftype = entity.dxftype()
            
            if dxftype not in ('LINE', 'LWPOLYLINE', 'POLYLINE', '3DPOLYLINE'):
                continue
            
            if entity.dxf.layer not in layers_set:
                continue
            
            # ============================================================
            # 1. LÍNEAS (LINE)
            # ============================================================
            if dxftype == 'LINE':
                s = entity.dxf.start
                e = entity.dxf.end
                lineas.append((
                    s.x, s.y, s.z if hasattr(s, 'z') else 0,
                    e.x, e.y, e.z if hasattr(e, 'z') else 0
                ))
                stats['LINE'] += 1
                stats['total_segmentos'] += 1
            
            # ============================================================
            # 2. POLILÍNEAS LIGERAS 2D (LWPOLYLINE)
            # ============================================================
            elif dxftype == 'LWPOLYLINE':
                elevation = entity.dxf.elevation if hasattr(entity.dxf, 'elevation') else 0
                pts = []
                
                # 🔧 FORMA CORRECTA: Usar context manager con 'with'
                try:
                    with entity.points() as p_list:
                        for p in p_list:
                            if len(p) >= 3:
                                pts.append((p[0], p[1], p[2]))
                            else:
                                pts.append((p[0], p[1], elevation))
                except Exception as e:
                    # Fallback: método alternativo
                    try:
                        for p in entity.points():
                            if len(p) >= 3:
                                pts.append((p[0], p[1], p[2]))
                            else:
                                pts.append((p[0], p[1], elevation))
                    except Exception as e2:
                        self.log_message(f"Error extrayendo puntos de LWPOLYLINE: {e2}", "WARNING")
                        continue
                
                if pts:
                    for i in range(len(pts) - 1):
                        lineas.append((
                            pts[i][0], pts[i][1], pts[i][2] if len(pts[i]) > 2 else elevation,
                            pts[i+1][0], pts[i+1][1], pts[i+1][2] if len(pts[i+1]) > 2 else elevation
                        ))
                        stats['total_segmentos'] += 1
                    
                    if hasattr(entity, 'closed') and entity.closed and len(pts) > 2:
                        lineas.append((
                            pts[-1][0], pts[-1][1], pts[-1][2] if len(pts[-1]) > 2 else elevation,
                            pts[0][0], pts[0][1], pts[0][2] if len(pts[0]) > 2 else elevation
                        ))
                        stats['total_segmentos'] += 1
                    
                    stats['LWPOLYLINE'] += 1
            
            # ============================================================
            # 3. POLILÍNEAS 3D (POLYLINE / 3DPOLYLINE)
            # ============================================================
            elif dxftype in ('POLYLINE', '3DPOLYLINE'):
                pts = []
                
                # 🔧 FORMA CORRECTA: Usar context manager con 'with'
                try:
                    with entity.points() as p_list:
                        for p in p_list:
                            if len(p) >= 3:
                                pts.append((p[0], p[1], p[2]))
                            else:
                                pts.append((p[0], p[1], 0))
                except Exception as e:
                    # Fallback: intentar con vertices() para POLYLINE 3D
                    try:
                        for vertex in entity.vertices():
                            loc = vertex.dxf.location
                            pts.append((loc.x, loc.y, loc.z))
                    except Exception as e2:
                        self.log_message(f"Error extrayendo puntos de POLYLINE: {e2}", "WARNING")
                        continue
                
                if pts:
                    for i in range(len(pts) - 1):
                        lineas.append((
                            pts[i][0], pts[i][1], pts[i][2] if len(pts[i]) > 2 else 0,
                            pts[i+1][0], pts[i+1][1], pts[i+1][2] if len(pts[i+1]) > 2 else 0
                        ))
                        stats['total_segmentos'] += 1
                    
                    if hasattr(entity, 'is_closed') and entity.is_closed and len(pts) > 2:
                        lineas.append((
                            pts[-1][0], pts[-1][1], pts[-1][2] if len(pts[-1]) > 2 else 0,
                            pts[0][0], pts[0][1], pts[0][2] if len(pts[0]) > 2 else 0
                        ))
                        stats['total_segmentos'] += 1
                    
                    stats['POLYLINE' if dxftype == 'POLYLINE' else '3DPOLYLINE'] += 1
        
        self.log_message(f"Extracción completada:", "INFO")
        self.log_message(f"   📊 Líneas (LINE): {stats['LINE']}", "INFO")
        self.log_message(f"   📊 Polilíneas 2D (LWPOLYLINE): {stats['LWPOLYLINE']}", "INFO")
        self.log_message(f"   📊 Polilíneas (POLYLINE): {stats['POLYLINE']}", "INFO")
        self.log_message(f"   📊 Polilíneas 3D (3DPOLYLINE): {stats['3DPOLYLINE']}", "INFO")
        self.log_message(f"   📐 Segmentos totales: {stats['total_segmentos']}", "INFO")
        
        return lineas

def normalizar_coordenadas(lineas, factor_conversion=1000):
    """Normaliza coordenadas al sistema de metros"""
    if not lineas:
        return lineas
    
    todos_puntos = []
    for l in lineas:
        todos_puntos.extend([(l[0], l[1]), (l[3], l[4])])
    
    min_x = min(p[0] for p in todos_puntos)
    min_y = min(p[1] for p in todos_puntos)
    
    lineas_norm = []
    for x1, y1, z1, x2, y2, z2 in lineas:
        lineas_norm.append((
            (x1 - min_x) / factor_conversion,
            (y1 - min_y) / factor_conversion,
            z1 / factor_conversion,
            (x2 - min_x) / factor_conversion,
            (y2 - min_y) / factor_conversion,
            z2 / factor_conversion
        ))
    
    return lineas_norm

def construir_red(lineas):
    """Construye la red a partir de las líneas normalizadas"""
    nodos_dict = {}
    tuberias = []
    next_id_nodo, next_id_tubo = 0, 0
    
    diametro_inicial = min(DIAMETROS_PAVCO.values())
    
    def get_nodo(x, y, z):
        nonlocal next_id_nodo
        for (xp, yp, zp), nid in nodos_dict.items():
            if abs(x-xp)<TOLERANCIA_CONEXION_M and abs(y-yp)<TOLERANCIA_CONEXION_M and abs(z-zp)<TOLERANCIA_CONEXION_M:
                return nid
        nodos_dict[(x, y, z)] = next_id_nodo
        next_id_nodo += 1
        return next_id_nodo - 1
    
    for x1,y1,z1,x2,y2,z2 in lineas:
        n1, n2 = get_nodo(x1, y1, z1), get_nodo(x2, y2, z2)
        L = np.sqrt((x2-x1)**2 + (y2-y1)**2 + (z2-z1)**2)
        tuberias.append(Tuberia(id=next_id_tubo, nodo_inicio=n1, nodo_fin=n2, longitud_m=L, diametro_mm=diametro_inicial))
        next_id_tubo += 1
    return nodos_dict, tuberias

def ajustar_cotas_relativas(red):
    if red.nodo_entrada_id is None:
        return
    
    z_entrada = red.nodos[red.nodo_entrada_id].z
    for nodo in red.nodos.values():
        nodo.z = nodo.z - z_entrada

# ================================================================================
# FUNCIONES DE VISUALIZACIÓN
# ================================================================================
def generate_3d_plot(red, presion_entrada=15.0):
    """Genera gráfico 3D completo con Plotly"""
    vel_min = st.session_state.get('vel_min', 0.5)
    vel_max = st.session_state.get('vel_max', 2.0)
    
    fig = go.Figure()
    
    diametros = [t.diametro_mm for t in red.tuberias.values() if t.diametro_mm > 0]
    if diametros:
        diam_min, diam_max = min(diametros), max(diametros)
    else:
        diam_min, diam_max = 21, 114
    
    for t in red.tuberias.values():
        if t.nodo_inicio not in red.nodos or t.nodo_fin not in red.nodos:
            continue
        n1, n2 = red.nodos[t.nodo_inicio], red.nodos[t.nodo_fin]
        v = t.velocidad_ms
        
        if v < vel_min:
            color, estado = '#74b9ff', 'Velocidad baja'
        elif v <= vel_max:
            color, estado = '#55efc4', 'Velocidad normal'
        else:
            color, estado = '#ff7675', 'Velocidad alta'
        
        if diam_max > diam_min:
            ancho = 4 + (t.diametro_mm - diam_min) / (diam_max - diam_min) * 6
        else:
            ancho = 6
        ancho = max(4, min(10, ancho))
        
        fig.add_trace(go.Scatter3d(
            x=[n1.x, n2.x], y=[n1.y, n2.y], z=[n1.z, n2.z],
            mode='lines',
            line=dict(color=color, width=ancho),
            name=f'Tubería {t.id}',
            showlegend=False,
            hovertext=f"<b>Tubería {t.id}</b><br>"
                     f"📏 Longitud: {t.longitud_m:.2f} m<br>"
                     f"⚙️ Diámetro: {t.diametro_mm:.0f} mm ({t.diametro_nominal_pulg}\")<br>"
                     f"💧 Caudal: {t.caudal_lps:.2f} L/s<br>"
                     f"⚡ Velocidad: {t.velocidad_ms:.2f} m/s ({estado})<br>"
                     f"📉 Pérdida: {t.perdida_mca:.2f} mca",
            hoverinfo='text'
        ))
    
    # Accesorios
    tipos_accesorios = {}
    for acc in red.accesorios:
        if acc.nodo_id not in red.nodos:
            continue
        if acc.tipo not in tipos_accesorios:
            tipos_accesorios[acc.tipo] = []
        tipos_accesorios[acc.tipo].append(acc)
    
    config_tipos = {
        'Tee': {'simbolo': 'square', 'color': '#f39c12', 'nombre': '🔧 Tee'},
        'Codo_90': {'simbolo': 'diamond', 'color': '#9b59b6', 'nombre': '🔀 Codo 90°'},
        'Codo_45': {'simbolo': 'diamond', 'color': '#8e44ad', 'nombre': '🔄 Codo 45°'},
        'Reduccion': {'simbolo': 'circle', 'color': '#e67e22', 'nombre': '📏 Reducción'},
    }
    
    for tipo, accesorios in tipos_accesorios.items():
        if tipo not in config_tipos:
            if 'Valvula' in tipo:
                nombre_valvula = tipo.split("_")[1]
                nombres_valvulas = {
                    'Compuerta': 'Válvula Compuerta',
                    'Globo': 'Válvula Globo', 
                    'Check': 'Válvula Check',
                    'Esfera': 'Válvula Esfera'
                }
                nombre = f'🔧 {nombres_valvulas.get(nombre_valvula, nombre_valvula)}'
                simbolo = 'diamond'
                color = '#e67e22'
            else:
                continue
        else:
            cfg = config_tipos[tipo]
            nombre = cfg['nombre']
            simbolo = cfg['simbolo']
            color = cfg['color']
        
        xs = [red.nodos[acc.nodo_id].x for acc in accesorios]
        ys = [red.nodos[acc.nodo_id].y for acc in accesorios]
        zs = [red.nodos[acc.nodo_id].z for acc in accesorios]
        
        hover_texts = []
        for acc in accesorios:
            nodo = red.nodos[acc.nodo_id]
            hover_texts.append(
                f"<b>{nombre}</b><br>📍 Nodo {acc.nodo_id}<br>"
                f"📐 Leq: {acc.longitud_equivalente_m:.2f} m<br>"
                f"📉 Pérdida: {acc.perdida_mca:.4f} mca"
            )
        
        fig.add_trace(go.Scatter3d(
            x=xs, y=ys, z=zs,
            mode='markers',
            marker=dict(size=7, symbol=simbolo, color=color, line=dict(width=1, color='white')),
            name=nombre,
            showlegend=True,
            hovertext=hover_texts,
            hoverinfo='text'
        ))
    
    # Nodos
    alcanzables = set()
    if red.nodo_entrada_id is not None:
        cola = deque([red.nodo_entrada_id])
        while cola:
            nodo_actual = cola.popleft()
            if nodo_actual in alcanzables:
                continue
            alcanzables.add(nodo_actual)
            
            nodo_obj = red.nodos.get(nodo_actual)
            if nodo_obj and nodo_obj.valvula_cerrada:
                continue
            
            for vecino in red.grafo.neighbors(nodo_actual):
                if vecino not in alcanzables:
                    cola.append(vecino)
    
    ug_acumulada = red.calcular_ug_acumulada(alcanzables=alcanzables)
    tipo_ocupacion = st.session_state.get('tipo_ocupacion', "Vivienda Unifamiliar")
    
    xs, ys, zs, colores, textos = [], [], [], [], []
    
    for n in red.nodos.values():
        xs.append(n.x)
        ys.append(n.y)
        zs.append(n.z)
        colores.append(n.presion_mca if n.presion_mca is not None else 0)
        
        demanda_unitaria = 0
        if n.tipo_aparato and n.tipo_aparato in UNIDADES_GASTO:
            demanda_unitaria = UNIDADES_GASTO[n.tipo_aparato]["caudal_unitario"]
        
        if n.id in alcanzables:
            ug_nodo = ug_acumulada.get(n.id, 0)
            caudal_hunter = caudal_por_ug(ug_nodo, tipo_ocupacion)
            caudal_texto = f"{caudal_hunter:.3f} L/s"
        else:
            caudal_texto = "N/A (Aislado por válvula)"
        
        z_rel = n.z
        z_signo = '+' if z_rel > 0 else ''
        
        texto = f"<b>Nodo {n.id}</b><br>"
        texto += f"📍 ({n.x:.2f}, {n.y:.2f}, {z_signo}{z_rel:.2f})<br>"
        
        if n.es_entrada:
            texto += "🚰 <b>ENTRADA (Cota 0)</b><br>"
        
        if n.tipo_aparato:
            ug = UNIDADES_GASTO.get(n.tipo_aparato, {}).get("ug", 0)
            texto += f"📌 {n.tipo_aparato} (UG: {ug})<br>"
            texto += f"💧 Demanda unitaria: {demanda_unitaria:.2f} L/s<br>"
        
        if n.valvula_tipo:
            estado_valvula = "CERRADA" if n.valvula_cerrada else "ABIERTA"
            texto += f"🔧 Válvula: {n.valvula_tipo} ({estado_valvula})<br>"
        
        texto += f"📊 Caudal Hunter en nodo: {caudal_texto}<br>"
        texto += f"💪 Presión: {n.presion_mca:.2f} mca" if n.presion_mca else "💪 Presión: N/A"
        
        if n.presion_mca and n.presion_mca < PRESION_MIN_NORMA:
            texto += "\n⚠️ <b>PRESIÓN BAJA</b>"
        
        textos.append(texto)
    
    fig.add_trace(go.Scatter3d(
        x=xs, y=ys, z=zs,
        mode='markers',
        marker=dict(
            size=8,
            color=colores,
            colorscale='RdYlGn',
            colorbar=dict(title="Presión (mca)", x=0.85, len=0.5, thickness=15),
            showscale=True,
            line=dict(width=1, color='white'),
            cmin=0,
            cmax=presion_entrada
        ),
        text=textos,
        hoverinfo='text',
        name='Nodos',
        showlegend=True
    ))
    
    fig.update_layout(
        title=dict(
            text="💧 Hydro Domus Py - Resultados del Análisis Hidráulico<br>"
                 "<sup>Cotas relativas al nodo de entrada (cota = 0)</sup>",
            font=dict(size=16),
            x=0.5,
            xanchor='center'
        ),
        scene=dict(
            xaxis_title="X (m)",
            yaxis_title="Y (m)",
            zaxis_title=dict(text="Altura relativa (m)", font=dict(size=12)),
            aspectmode='data',
            camera=dict(eye=dict(x=1.5, y=1.5, z=1.5)),
        ),
        width=1200,
        height=700,
        legend=dict(
            x=0.02,
            y=0.98,
            traceorder='normal'
        ),
        margin=dict(l=0, r=20, t=70, b=0),
        hoverlabel=dict(bgcolor='white', font_size=11)
    )
    
    return fig

def generar_perfil_presiones(red):
    """Genera perfil de presiones 2D"""
    try:
        if red.nodo_entrada_id is None:
            return None
        
        distancias = nx.single_source_dijkstra_path_length(red.grafo, red.nodo_entrada_id)
        if distancias:
            nodo_lejano = max(distancias, key=distancias.get)
            camino = nx.shortest_path(red.grafo, red.nodo_entrada_id, nodo_lejano)
        else:
            camino = list(red.nodos.keys())[:10]
        
        distancias_acum = [0]
        presiones = [red.nodos[camino[0]].presion_mca or 0]
        cotas = [red.nodos[camino[0]].z]
        
        for i in range(1, len(camino)):
            tid = red._find_tuberia(camino[i-1], camino[i])
            if tid:
                dist_acum = distancias_acum[-1] + red.tuberias[tid].longitud_m
            else:
                dist_acum = distancias_acum[-1]
            distancias_acum.append(dist_acum)
            presiones.append(red.nodos[camino[i]].presion_mca or 0)
            cotas.append(red.nodos[camino[i]].z)
        
        fig = make_subplots(rows=2, cols=1, shared_xaxes=True, vertical_spacing=0.1,
                           subplot_titles=("📈 Perfil de Presión", "📍 Perfil de Cota"))
        
        fig.add_trace(go.Scatter(x=distancias_acum, y=presiones, mode='lines+markers', 
                                 name='Presión', line=dict(color='#3498db', width=3),
                                 marker=dict(size=8, color='#3498db')), row=1, col=1)
        
        fig.add_hline(y=PRESION_MIN_NORMA, line_dash="dash", line_color="#e74c3c", 
                      annotation_text=f"Presión mínima norma ({PRESION_MIN_NORMA} mca)", 
                      annotation_position="top right", row=1, col=1)
        
        presion_entrada = st.session_state.get('presion_entrada', 15.0)
        fig.add_hline(y=presion_entrada, line_dash="dot", line_color="#2ecc71",
                      annotation_text=f"Presión entrada ({presion_entrada} mca)",
                      annotation_position="bottom right", row=1, col=1)
        
        fig.add_trace(go.Scatter(x=distancias_acum, y=cotas, mode='lines+markers', 
                                 name='Cota', line=dict(color='#e67e22', width=3),
                                 marker=dict(size=8, color='#e67e22')), row=2, col=1)
        
        fig.update_layout(
            title=dict(
                text="💧 Perfil Hidráulico - Ruta más larga desde la entrada",
                font=dict(size=16),
                x=0.5
            ),
            height=500,
            template='plotly_white'
        )
        
        fig.update_xaxes(title_text="Distancia acumulada (m)", row=2, col=1)
        fig.update_yaxes(title_text="Presión (mca)", row=1, col=1)
        fig.update_yaxes(title_text="Cota (m)", row=2, col=1)
        
        return fig
    except Exception as e:
        return None

def generar_reporte_materiales(red):
    """Genera reporte de materiales en formato DataFrame"""
    longitudes = {}
    for t in red.tuberias.values():
        diam = t.diametro_nominal_pulg
        longitudes[diam] = longitudes.get(diam, 0) + t.longitud_m
    
    tramos = {d: int(np.ceil(l / 6.0)) for d, l in longitudes.items()}
    
    accesorios = {}
    for acc in red.accesorios:
        nombre = acc.tipo.replace("_", " ")
        if "Valvula" in nombre:
            nombre = nombre.replace("Valvula ", "Válvula ")
        accesorios[nombre] = accesorios.get(nombre, 0) + 1
    
    df_tuberias = pd.DataFrame([
        {
            "Diámetro": d,
            "DI (mm)": DIAMETROS_PAVCO.get(d.replace("-", "_"), 0),
            "Longitud total (m)": round(l, 2),
            "Tramos de 6m": tramos[d]
        }
        for d, l in sorted(longitudes.items(), key=lambda x: diametro_a_numero(x[0]))
    ])
    
    df_accesorios = pd.DataFrame([
        {"Tipo": t, "Cantidad": c}
        for t, c in sorted(accesorios.items(), key=lambda x: -x[1])
    ])
    
    total_long = sum(longitudes.values())
    total_tramos = sum(tramos.values())
    total_acc = sum(accesorios.values())
    
    return df_tuberias, df_accesorios, total_long, total_tramos, total_acc

# ================================================================================
# FUNCIONES DE EXPORTACIÓN
# ================================================================================
def exportar_excel(red):
    """Exporta resultados a Excel"""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    
    output = io.BytesIO()
    wb = openpyxl.Workbook()
    
    default_sheet = wb.active
    wb.remove(default_sheet)
    
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="FF2C3E50", end_color="FF2C3E50", fill_type="solid")
    title_font = Font(bold=True, size=14, color="FF1A5276")
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                         top=Side(style='thin'), bottom=Side(style='thin'))
    green_fill = PatternFill(start_color="FF27AE60", end_color="FF27AE60", fill_type="solid")
    red_fill = PatternFill(start_color="FFE74C3C", end_color="FFE74C3C", fill_type="solid")
    section_fill = PatternFill(start_color="FFD5DBDB", end_color="FFD5DBDB", fill_type="solid")
    
    # HOJA 1: RESUMEN
    ws_resumen = wb.create_sheet("1. Resumen Ejecutivo", 0)
    
    presiones = [n.presion_mca for n in red.nodos.values() if n.presion_mca is not None]
    ug_acumulada = red.calcular_ug_acumulada()
    ug_total = ug_acumulada.get(red.nodo_entrada_id, 0)
    tipo_ocupacion = st.session_state.get('tipo_ocupacion', "Vivienda Unifamiliar")
    caudal_total = caudal_por_ug(ug_total, tipo_ocupacion)
    presion_entrada = st.session_state.get('presion_entrada', 15.0)
    cumple = min(presiones) >= PRESION_MIN_NORMA if presiones else False
    
    aparatos = sum(1 for n in red.nodos.values() if n.tipo_aparato)
    valvulas = sum(1 for n in red.nodos.values() if n.valvula_tipo)
    valvulas_cerradas = sum(1 for n in red.nodos.values() if n.valvula_cerrada)
    
    ws_resumen.merge_cells('A1:D1')
    cell = ws_resumen.cell(row=1, column=1, value="HYDRO DOMUS PY - ANÁLISIS HIDRÁULICO")
    cell.font = Font(bold=True, size=16, color="FF1A5276")
    cell.alignment = Alignment(horizontal="center")
    
    ws_resumen.merge_cells('A2:D2')
    cell = ws_resumen.cell(row=2, column=1, value=f"Fecha: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    cell.alignment = Alignment(horizontal="center")
    
    row = 4
    
    ws_resumen.merge_cells(f'A{row}:D{row}')
    cell = ws_resumen.cell(row=row, column=1, value="📐 PARÁMETROS DE CÁLCULO")
    cell.font = title_font
    cell.fill = section_fill
    row += 1
    
    params_data = [
        ["Tipo de edificación", tipo_ocupacion],
        ["Presión disponible en entrada", f"{presion_entrada} mca"],
        ["Velocidad mínima", f"{st.session_state.get('vel_min', 0.5)} m/s"],
        ["Velocidad máxima", f"{st.session_state.get('vel_max', 2.0)} m/s"],
        ["Presión mínima requerida (NTC 1500)", f"{PRESION_MIN_NORMA} mca"],
    ]
    
    for param in params_data:
        ws_resumen.cell(row=row, column=1, value=param[0]).font = Font(bold=True)
        ws_resumen.cell(row=row, column=2, value=param[1])
        ws_resumen.cell(row=row, column=1).border = thin_border
        ws_resumen.cell(row=row, column=2).border = thin_border
        row += 1
    
    row += 1
    
    ws_resumen.merge_cells(f'A{row}:D{row}')
    cell = ws_resumen.cell(row=row, column=1, value="📊 ESTADÍSTICAS DE LA RED")
    cell.font = title_font
    cell.fill = section_fill
    row += 1
    
    stats_data = [
        ["Nodos totales", len(red.nodos)],
        ["Tuberías", len(red.tuberias)],
        ["Accesorios detectados", len(red.accesorios)],
        ["Aparatos sanitarios asignados", aparatos],
        ["Válvulas totales", valvulas],
        ["Válvulas cerradas", valvulas_cerradas],
    ]
    
    for stat in stats_data:
        ws_resumen.cell(row=row, column=1, value=stat[0]).font = Font(bold=True)
        ws_resumen.cell(row=row, column=2, value=stat[1])
        ws_resumen.cell(row=row, column=1).border = thin_border
        ws_resumen.cell(row=row, column=2).border = thin_border
        row += 1
    
    row += 1
    
    ws_resumen.merge_cells(f'A{row}:D{row}')
    cell = ws_resumen.cell(row=row, column=1, value="💧 CAUDALES")
    cell.font = title_font
    cell.fill = section_fill
    row += 1
    
    caudal_data = [
        ["Unidades de gasto totales (UG)", f"{ug_total:.0f}"],
        ["Caudal máximo probable (Hunter)", f"{caudal_total:.2f} L/s"],
        ["Fórmula aplicada", TIPOS_OCUPACION_AGUA.get(tipo_ocupacion, {}).get("formula", "0.2√UG + 0.5")],
    ]
    
    for caud in caudal_data:
        ws_resumen.cell(row=row, column=1, value=caud[0]).font = Font(bold=True)
        ws_resumen.cell(row=row, column=2, value=caud[1])
        ws_resumen.cell(row=row, column=1).border = thin_border
        ws_resumen.cell(row=row, column=2).border = thin_border
        row += 1
    
    row += 1
    
    ws_resumen.merge_cells(f'A{row}:D{row}')
    cell = ws_resumen.cell(row=row, column=1, value="💪 PRESIONES")
    cell.font = title_font
    cell.fill = section_fill
    row += 1
    
    if presiones:
        p_min, p_max = min(presiones), max(presiones)
        p_prom = sum(presiones) / len(presiones)
        presion_data = [
            ["Presión en entrada", f"{presion_entrada:.1f} mca"],
            ["Presión mínima", f"{p_min:.2f} mca"],
            ["Presión máxima", f"{p_max:.2f} mca"],
            ["Presión promedio", f"{p_prom:.2f} mca"],
        ]
        
        for pres in presion_data:
            ws_resumen.cell(row=row, column=1, value=pres[0]).font = Font(bold=True)
            ws_resumen.cell(row=row, column=2, value=pres[1])
            ws_resumen.cell(row=row, column=1).border = thin_border
            ws_resumen.cell(row=row, column=2).border = thin_border
            row += 1
        
        row += 1
        
        ws_resumen.merge_cells(f'A{row}:D{row}')
        if cumple:
            cell = ws_resumen.cell(row=row, column=1, value="✅ CUMPLIMIENTO NORMA NTC 1500")
            cell.font = Font(bold=True, size=12, color="FFFFFF")
            cell.fill = green_fill
            cell.alignment = Alignment(horizontal="center")
        else:
            cell = ws_resumen.cell(row=row, column=1, value="❌ NO CUMPLE NORMA NTC 1500")
            cell.font = Font(bold=True, size=12, color="FFFFFF")
            cell.fill = red_fill
            cell.alignment = Alignment(horizontal="center")
    
    ws_resumen.column_dimensions['A'].width = 35
    ws_resumen.column_dimensions['B'].width = 25
    
    # HOJA 2: NODOS
    ws_nodos = wb.create_sheet("2. Nodos", 1)
    headers_nodos = ["ID", "X (m)", "Y (m)", "Z (m)", "Presión (mca)", "Aparato", "Válvula", "Entrada"]
    for col, header in enumerate(headers_nodos, 1):
        cell = ws_nodos.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border
    
    for row_idx, n in enumerate(red.nodos.values(), 2):
        valvula_texto = f"{n.valvula_tipo} ({'CERRADA' if n.valvula_cerrada else 'ABIERTA'})" if n.valvula_tipo else "-"
        ws_nodos.cell(row=row_idx, column=1, value=n.id).border = thin_border
        ws_nodos.cell(row=row_idx, column=2, value=round(n.x, 2)).border = thin_border
        ws_nodos.cell(row=row_idx, column=3, value=round(n.y, 2)).border = thin_border
        ws_nodos.cell(row=row_idx, column=4, value=round(n.z, 2)).border = thin_border
        ws_nodos.cell(row=row_idx, column=5, value=round(n.presion_mca, 2) if n.presion_mca else None).border = thin_border
        ws_nodos.cell(row=row_idx, column=6, value=n.tipo_aparato or "-").border = thin_border
        ws_nodos.cell(row=row_idx, column=7, value=valvula_texto).border = thin_border
        ws_nodos.cell(row=row_idx, column=8, value="✓" if n.es_entrada else "").border = thin_border
    
    for col in range(1, len(headers_nodos) + 1):
        ws_nodos.column_dimensions[get_column_letter(col)].width = 12
    
    # HOJA 3: TUBERÍAS
    ws_tuberias = wb.create_sheet("3. Tuberías", 2)
    headers_tuberias = ["ID", "Desde", "Hasta", "Longitud (m)", "Diámetro", "DI (mm)", "Caudal (L/s)", "Velocidad (m/s)", "Pérdida (mca)"]
    for col, header in enumerate(headers_tuberias, 1):
        cell = ws_tuberias.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border
    
    for row_idx, t in enumerate(red.tuberias.values(), 2):
        ws_tuberias.cell(row=row_idx, column=1, value=t.id).border = thin_border
        ws_tuberias.cell(row=row_idx, column=2, value=t.nodo_inicio).border = thin_border
        ws_tuberias.cell(row=row_idx, column=3, value=t.nodo_fin).border = thin_border
        ws_tuberias.cell(row=row_idx, column=4, value=round(t.longitud_m, 2)).border = thin_border
        ws_tuberias.cell(row=row_idx, column=5, value=t.diametro_nominal_pulg).border = thin_border
        ws_tuberias.cell(row=row_idx, column=6, value=round(t.diametro_mm, 2)).border = thin_border
        ws_tuberias.cell(row=row_idx, column=7, value=round(t.caudal_lps, 3)).border = thin_border
        ws_tuberias.cell(row=row_idx, column=8, value=round(t.velocidad_ms, 2)).border = thin_border
        ws_tuberias.cell(row=row_idx, column=9, value=round(t.perdida_mca, 3)).border = thin_border
    
    for col in range(1, len(headers_tuberias) + 1):
        ws_tuberias.column_dimensions[get_column_letter(col)].width = 12
    
    # HOJA 4: ACCESORIOS
    ws_accesorios = wb.create_sheet("4. Accesorios", 3)
    headers_accesorios = ["ID", "Tipo", "Nodo", "DN (pulg)", "DI (mm)", "Leq (m)", "Pérdida (mca)"]
    for col, header in enumerate(headers_accesorios, 1):
        cell = ws_accesorios.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border
    
    for row_idx, acc in enumerate(red.accesorios, 2):
        diametro_nom = "N/A"
        diametro_mm = 0
        for t in red.tuberias.values():
            if t.nodo_inicio == acc.nodo_id or t.nodo_fin == acc.nodo_id:
                diametro_nom = t.diametro_nominal_pulg
                diam_key = diametro_nom.replace("-", "_")
                diametro_mm = DIAMETROS_PAVCO.get(diam_key, 0)
                break
        
        nombre = acc.tipo.replace("_", " ")
        if "Valvula" in nombre:
            nombre = nombre.replace("Valvula ", "Válvula ")
        
        ws_accesorios.cell(row=row_idx, column=1, value=acc.id).border = thin_border
        ws_accesorios.cell(row=row_idx, column=2, value=nombre).border = thin_border
        ws_accesorios.cell(row=row_idx, column=3, value=acc.nodo_id).border = thin_border
        ws_accesorios.cell(row=row_idx, column=4, value=diametro_nom).border = thin_border
        ws_accesorios.cell(row=row_idx, column=5, value=round(diametro_mm, 2)).border = thin_border
        ws_accesorios.cell(row=row_idx, column=6, value=round(acc.longitud_equivalente_m, 2)).border = thin_border
        ws_accesorios.cell(row=row_idx, column=7, value=round(acc.perdida_mca, 4)).border = thin_border
    
    for col in range(1, len(headers_accesorios) + 1):
        ws_accesorios.column_dimensions[get_column_letter(col)].width = 14
    
    wb.save(output)
    return output.getvalue()

# ================================================================================
# CONFIGURACIÓN INTERACTIVA (HTML + JavaScript embebido)
# ================================================================================
def generar_html_config(nodos_data, tuberias_data):
    """Genera HTML para la configuración interactiva de nodos"""
    
    html = f"""<!DOCTYPE html>
<html><head><meta charset="UTF-8"><title>Hydro Domus Py - Configuración</title>
<script src="https://cdn.plot.ly/plotly-3.0.1.min.js"></script>
<style>
body{{margin:0;font-family:Segoe UI,sans-serif;background:#1e1e1e;color:#ffffff}}
#header{{background:#2c3e50;color:white;padding:15px;text-align:center}}
#header h1{{margin:0;font-size:1.5em}}
#header p{{margin:5px 0 0;font-size:0.9em}}
#main{{position:fixed;top:80px;left:0;right:200px;bottom:0}}
#sidebar{{position:fixed;top:80px;right:0;width:200px;bottom:0;background:#2d2d2d;padding:12px;overflow-y:auto;font-size:11px;box-shadow:-2px 0 10px rgba(0,0,0,0.3)}}
#sidebar h3{{margin-top:0;color:#3498db;font-size:14px}}
.btn{{background:#3498db;color:white;border:none;padding:6px 10px;border-radius:6px;cursor:pointer;margin:4px;font-size:11px;transition:all 0.3s}}
.btn:hover{{background:#2980b9;transform:scale(1.02)}}
.btn-success{{background:#27ae60}}
.btn-success:hover{{background:#229954}}
.btn-danger{{background:#e74c3c}}
.btn-danger:hover{{background:#c0392b}}
.badge-entrada{{background:#e74c3c;color:white;padding:2px 8px;border-radius:12px;font-size:10px;display:inline-block;margin:2px}}
.badge-aparato{{background:#3498db;color:white;padding:2px 8px;border-radius:12px;font-size:10px;display:inline-block;margin:2px}}
.badge-valvula{{background:#e67e22;color:white;padding:2px 8px;border-radius:12px;font-size:10px;display:inline-block;margin:2px}}
.info-nodo{{background:#3d3d3d;border-radius:8px;padding:10px;margin-bottom:10px;border:1px solid #444444;color:#ffffff}}
.info-nodo h4{{margin:0 0 6px 0;color:#3498db;font-size:12px}}
.info-nodo p{{margin:4px 0;font-size:10px;color:#aaaaaa}}
select{{font-size:10px;padding:3px;margin:3px 0;width:100%;border-radius:4px;border:1px solid #444444;background:#2d2d2d;color:#ffffff}}
hr{{margin:8px 0;border-color:#444444}}
label{{color:#ffffff}}
</style>
</head>
<body>
<div id="header"><h1>💧 Hydro Domus Py - Configuración Interactiva</h1><p>Haga clic en cualquier nodo para configurarlo</p></div>
<div id="main"></div>
<div id="sidebar">
<h3>📋 Configuración</h3>
<div id="info-panel"><p style="font-size:11px; color:#aaaaaa">✨ Haga clic en un nodo del gráfico 3D</p></div>
<div style="text-align:center; margin-top:10px">
<button class="btn btn-success" onclick="guardar()" style="width:95%">💾 Guardar Configuración</button>
<button class="btn btn-danger" onclick="resetear()" style="width:95%; margin-top:6px">🔄 Resetear todo</button>
</div>
<hr>
<div id="resumen" style="font-size:11px; background:#3d3d3d; padding:8px; border-radius:6px;"></div>
<p style="font-size:8px; color:#999; text-align:center; margin-top:10px">Hydro Domus Py v2.1.0<br>NTC 1500 - RAS</p>
</div>
<script>
const nodos = {json.dumps(nodos_data)};
const tuberias = {json.dumps(tuberias_data)};
const aparatos = {json.dumps(list(UNIDADES_GASTO.keys()))};
let entradaId = null;
let nodoSeleccionado = null;
let currentCamera = null;

function actualizarGrafico() {{
    const tx=[],ty=[],tz=[];
    for(const t of tuberias){{tx.push(t.x1,t.x2,null);ty.push(t.y1,t.y2,null);tz.push(t.z1,t.z2,null);}}
    const nx=[],ny=[],nz=[],col=[],tam=[];
    for(const n of nodos){{
        nx.push(n.x);ny.push(n.y);nz.push(n.z);
        if(n.id===entradaId){{col.push('#e74c3c');tam.push(12);}}
        else if(n.tipo_aparato){{col.push('#3498db');tam.push(9);}}
        else if(n.valvula_tipo){{col.push('#e67e22');tam.push(9);}}
        else{{col.push('#95a5a6');tam.push(6);}}
    }}
    
    var xs = nodos.map(n => n.x);
    var ys = nodos.map(n => n.y);
    var zs = nodos.map(n => n.z);
    var xRange = Math.max(...xs) - Math.min(...xs);
    var yRange = Math.max(...ys) - Math.min(...ys);
    var zRange = Math.max(...zs) - Math.min(...zs);
    var maxRange = Math.max(xRange, yRange, zRange);
    
    const layout = {{
        scene: {{
            xaxis: {{title: 'X (m)'}},
            yaxis: {{title: 'Y (m)'}},
            zaxis: {{title: 'Z (m)'}},
            aspectmode: 'manual',
            aspectratio: {{x: xRange / maxRange, y: yRange / maxRange, z: zRange / maxRange}},
            camera: currentCamera || {{eye: {{x: 1.5, y: 1.5, z: 1.5}}}},
            bgcolor: '#1e1e1e'
        }},
        margin: {{l: 0, r: 0, t: 40, b: 0}},
        paper_bgcolor: '#1e1e1e',
        plot_bgcolor: '#1e1e1e'
    }};
    
    const traces = [
        {{type:'scatter3d',mode:'lines',x:tx,y:ty,z:tz,line:{{color:'#bdc3c7',width:4}}}},
        {{type:'scatter3d',mode:'markers+text',x:nx,y:ny,z:nz,marker:{{color:col,size:tam}},
          text:nodos.map(n=>n.id), textposition: 'top center', textfont: {{size: 10, color:'#ffffff'}}, customdata: nodos.map(n=>n.id), hoverinfo: 'text'}}
    ];
    
    Plotly.newPlot('main', traces, layout);
    
    document.getElementById('main').on('plotly_click',function(data){{
        if(data.points&&data.points[0]){{
            const nodo=nodos.find(n=>n.id===data.points[0].customdata);
            if(nodo)mostrarPanel(nodo);
        }}
    }});
}}

function mostrarPanel(nodo){{
    nodoSeleccionado=nodo;
    const esEntrada=(nodo.id===entradaId);
    const tieneAparato=!!nodo.tipo_aparato;
    const tieneValvula=!!nodo.valvula_tipo;
    const valvulaCerrada=nodo.valvula_cerrada || false;
    
    let html=`<div class="info-nodo"><h4>🔘 Nodo ${{nodo.id}}</h4>
    <p>📍 (${{nodo.x.toFixed(1)}}, ${{nodo.y.toFixed(1)}}, ${{nodo.z.toFixed(1)}})</p>
    <p>📌 Estado: `;
    if(esEntrada)html+='<span class="badge-entrada">🚰 ENTRADA</span>';
    if(tieneAparato)html+=`<span class="badge-aparato">📌 ${{nodo.tipo_aparato}}</span>`;
    if(tieneValvula){{
        const estadoValvula = valvulaCerrada ? 'CERRADA' : 'ABIERTA';
        html+=`<span class="badge-valvula">🔧 ${{nodo.valvula_tipo}} (${{estadoValvula}})</span>`;
    }}
    if(!esEntrada&&!tieneAparato&&!tieneValvula) html+='⚪ Sin asignar';
    
    html+=`</p><hr>
    <button class="btn" onclick="setEntrada(${{nodo.id}})" ${{esEntrada?'disabled':''}} style="width:100%">🚰 Hacer ENTRADA</button>
    
    <div style="margin-top:5px"><label>📌 Aparato:</label>
    <select id="selAparato" style="width:100%"><option value="">-- Seleccionar --</option>`;
    for(const a of aparatos)html+=`<option value="${{a}}" ${{nodo.tipo_aparato===a?'selected':''}}>${{a}}</option>`;
    html+=`</select>
    <button class="btn" onclick="setAparato(${{nodo.id}})" style="width:100%">Aplicar</button></div>
    
    <div style="margin-top:5px"><label>🔧 Válvula:</label>
    <select id="selValvula" style="width:100%">
        <option value="">-- Ninguna --</option>
        <option value="Compuerta" ${{nodo.valvula_tipo==='Compuerta'?'selected':''}}>Compuerta (Leq=0.3m)</option>
        <option value="Globo" ${{nodo.valvula_tipo==='Globo'?'selected':''}}>Globo (Leq=1.5m)</option>
        <option value="Check" ${{nodo.valvula_tipo==='Check'?'selected':''}}>Check (Leq=2.5m)</option>
        <option value="Esfera" ${{nodo.valvula_tipo==='Esfera'?'selected':''}}>Esfera (Leq=0.2m)</option>
    </select>
    <select id="selEstadoValvula" style="width:100%; margin-top:2px" ${{!tieneValvula?'disabled':''}}>
        <option value="abierta" ${{!valvulaCerrada?'selected':''}}>ABIERTA (flujo normal)</option>
        <option value="cerrada" ${{valvulaCerrada?'selected':''}}>CERRADA (aisla aguas abajo)</option>
    </select>
    <button class="btn" onclick="setValvula(${{nodo.id}})" style="width:100%">Aplicar</button></div>`;
    
    if(!esEntrada&&!tieneAparato&&!tieneValvula)html+=`<hr><button class="btn" onclick="limpiarNodo(${{nodo.id}})" style="width:100%">🗑️ Limpiar todo</button>`;
    html+=`</div>`;
    document.getElementById('info-panel').innerHTML=html;
}}

function setEntrada(id){{
    if(entradaId){{
        const old=nodos.find(n=>n.id===entradaId);
        if(old)old.es_entrada=false;
    }}
    entradaId=id;
    const nodo=nodos.find(n=>n.id===id);
    if(nodo){{nodo.es_entrada=true;nodo.tipo_aparato="";nodo.valvula_tipo="";nodo.valvula_cerrada=false;}}
    actualizarGrafico();actualizarResumen();
    if(nodoSeleccionado&&nodoSeleccionado.id===id)mostrarPanel(nodo);
}}

function setAparato(id){{
    const tipo=document.getElementById('selAparato').value;
    if(!tipo)return;
    if(id===entradaId){{alert('La entrada no puede ser aparato');return;}}
    const nodo=nodos.find(n=>n.id===id);
    if(nodo){{nodo.tipo_aparato=tipo;nodo.es_entrada=false;}}
    actualizarGrafico();actualizarResumen();
    if(nodoSeleccionado&&nodoSeleccionado.id===id)mostrarPanel(nodo);
}}

function setValvula(id){{
    const tipo = document.getElementById('selValvula').value;
    const estado = document.getElementById('selEstadoValvula').value;
    const nodo = nodos.find(n=>n.id===id);
    if(!nodo) return;
    
    if(!tipo){{
        nodo.valvula_tipo = "";
        nodo.valvula_cerrada = false;
    }} else {{
        nodo.valvula_tipo = tipo;
        nodo.valvula_cerrada = (estado === 'cerrada');
    }}
    actualizarGrafico();actualizarResumen();
    if(nodoSeleccionado && nodoSeleccionado.id===id) mostrarPanel(nodo);
}}

function limpiarNodo(id){{
    const nodo=nodos.find(n=>n.id===id);
    if(nodo){{
        if(nodo.id===entradaId)entradaId=null;
        nodo.es_entrada=false;
        nodo.tipo_aparato="";
        nodo.valvula_tipo="";
        nodo.valvula_cerrada=false;
    }}
    actualizarGrafico();actualizarResumen();
    if(nodoSeleccionado&&nodoSeleccionado.id===id)mostrarPanel(nodo);
}}

function resetear(){{
    if(confirm('¿Resetear toda la configuración?')){{
        entradaId=null;
        for(const n of nodos){{
            n.es_entrada=false;
            n.tipo_aparato="";
            n.valvula_tipo="";
            n.valvula_cerrada=false;
        }}
        actualizarGrafico();actualizarResumen();
        document.getElementById('info-panel').innerHTML='<p style="font-size:11px; color:#aaaaaa">Seleccione un nodo</p>';
    }}
}}

function actualizarResumen(){{
    const entrada=nodos.find(n=>n.id===entradaId);
    const aparatosList=nodos.filter(n=>n.tipo_aparato);
    const valvulasList=nodos.filter(n=>n.valvula_tipo);
    const valvulasCerradas=valvulasList.filter(n=>n.valvula_cerrada);
    const totalUG = aparatosList.reduce((sum, n) => {{
        const ugMap = {json.dumps({k:v["ug"] for k,v in UNIDADES_GASTO.items()})};
        return sum + (ugMap[n.tipo_aparato] || 0);
    }}, 0);
    document.getElementById('resumen').innerHTML=`<strong>🚰 Entrada:</strong> ${{entrada?'Nodo '+entrada.id:'❌ No'}}<br>
    <strong>📌 Aparatos:</strong> ${{aparatosList.length}}<br>
    <strong>🔧 Válvulas:</strong> ${{valvulasList.length}} (${{valvulasCerradas.length}} cerradas)<br>
    <strong>📊 UG totales:</strong> ${{totalUG}} UG<br>
    <strong>💧 Caudal probable:</strong> ${{(0.2 * Math.sqrt(totalUG) + 0.5).toFixed(2)}} L/s`;
}}

function guardar(){{
    if(!entradaId){{alert('Seleccione un nodo de entrada');return;}}
    const config={{nodo_entrada:entradaId,nodos:nodos.map(n=>({{
        id:n.id,
        es_entrada:n.id===entradaId,
        tipo_aparato:n.tipo_aparato,
        valvula_tipo:n.valvula_tipo,
        valvula_cerrada:n.valvula_cerrada || false
    }}))}};
    const blob=new Blob([JSON.stringify(config,null,2)],{{type:'application/json'}});
    const a=document.createElement('a');
    a.href=URL.createObjectURL(blob);
    a.download='HydroDomusPy_config.json';
    a.click();
    alert('✅ Configuración guardada. Cierre esta ventana y continúe en la aplicación.');
}}

actualizarGrafico();actualizarResumen();
</script></body></html>"""
    return html

# ================================================================================
# INTERFAZ PRINCIPAL DE STREAMLIT
# ================================================================================
def main():
    """Función principal de la aplicación"""
    
    # ============================================================
    # INICIALIZAR ESTADO DE SESIÓN
    # ============================================================
    if 'red' not in st.session_state:
        st.session_state.red = None
    if 'analyzer' not in st.session_state:
        st.session_state.analyzer = None
    if 'presion_entrada' not in st.session_state:
        st.session_state.presion_entrada = 15.0
    if 'tipo_ocupacion' not in st.session_state:
        st.session_state.tipo_ocupacion = "Vivienda Unifamiliar"
    if 'diametro_maximo' not in st.session_state:
        st.session_state.diametro_maximo = None
    if 'resultados' not in st.session_state:
        st.session_state.resultados = None
    if 'config_aplicada' not in st.session_state:
        st.session_state.config_aplicada = False
    if 'log_messages' not in st.session_state:
        st.session_state.log_messages = []
    if 'dxf_loaded' not in st.session_state:
        st.session_state.dxf_loaded = False
    if 'dxf_layers' not in st.session_state:
        st.session_state.dxf_layers = []
    if 'selected_layers' not in st.session_state:
        st.session_state.selected_layers = []
    if 'vel_min' not in st.session_state:
        st.session_state.vel_min = 0.5
    if 'vel_max' not in st.session_state:
        st.session_state.vel_max = 2.0
    if 'tmp_dxf_path' not in st.session_state:
        st.session_state.tmp_dxf_path = None
    if 'dxf_reader' not in st.session_state:
        st.session_state.dxf_reader = None
    if 'unidad_dibujo' not in st.session_state:
        st.session_state.unidad_dibujo = "mm"
    if 'factor_conversion' not in st.session_state:
        st.session_state.factor_conversion = 1000
    if 'paso_actual' not in st.session_state:
        st.session_state.paso_actual = 1  # 1=DXF, 2=Configurar, 3=Analizar

    # ============================================================
    # BARRA LATERAL COMPACTA
    # ============================================================
    with st.sidebar:
        # Logo y título más compacto
        st.markdown("""
        <div style="text-align:center; padding:5px 0;">
            <h1 style="font-size:1.8em; margin:0;">💧</h1>
            <h2 style="font-size:1.2em; margin:0;">Hydro Domus Py</h2>
            <p style="font-size:0.8em; color:#888; margin:0;">Análisis Hidráulico v2.1</p>
        </div>
        """, unsafe_allow_html=True)
        
        st.divider()
        
        # ===== ASISTENTE PASO A PASO =====
        st.markdown("### 📋 Progreso")
        
        pasos = [
            ("📁", "Cargar DXF", 1),
            ("🔧", "Configurar", 2),
            ("🚀", "Analizar", 3),
        ]
        
        cols = st.columns(3)
        for i, (icon, label, paso) in enumerate(pasos):
            with cols[i]:
                if st.session_state.paso_actual >= paso:
                    st.markdown(f"""
                    <div style="text-align:center; background:#2d7d46; padding:5px; border-radius:8px;">
                        <span style="font-size:1.5em;">{icon}</span>
                        <p style="font-size:0.7em; margin:0; color:white;">{label}</p>
                    </div>
                    """, unsafe_allow_html=True)
                elif st.session_state.paso_actual == paso - 1:
                    st.markdown(f"""
                    <div style="text-align:center; background:#f39c12; padding:5px; border-radius:8px;">
                        <span style="font-size:1.5em;">{icon}</span>
                        <p style="font-size:0.7em; margin:0; color:white;">{label}</p>
                    </div>
                    """, unsafe_allow_html=True)
                else:
                    st.markdown(f"""
                    <div style="text-align:center; background:#444; padding:5px; border-radius:8px;">
                        <span style="font-size:1.5em; opacity:0.5;">{icon}</span>
                        <p style="font-size:0.7em; margin:0; color:#888;">{label}</p>
                    </div>
                    """, unsafe_allow_html=True)
        
        st.divider()
        
        # ===== PASO 1: CARGAR DXF =====
        with st.expander("📁 Cargar DXF", expanded=(st.session_state.paso_actual == 1)):
            dxf_file = st.file_uploader(
                "Seleccionar archivo DXF",
                type=['dxf'],
                help="Cargue el plano hidrosanitario en formato DXF",
                key="dxf_uploader_main"
            )
            
            # Unidades
            unidad_seleccionada = st.selectbox(
                "Unidades del dibujo:",
                options=list(UNIDADES_DIBUJO.keys()),
                format_func=lambda x: f"{UNIDADES_DIBUJO[x]['icono']} {UNIDADES_DIBUJO[x]['nombre']}",
                key="unidad_select_main"
            )
            st.session_state.unidad_dibujo = unidad_seleccionada
            st.session_state.factor_conversion = UNIDADES_DIBUJO[unidad_seleccionada]["factor"]
            
            # Procesar DXF
            if dxf_file is not None:
                st.success(f"✅ {dxf_file.name}")
                
                with tempfile.NamedTemporaryFile(delete=False, suffix='.dxf') as tmp:
                    tmp.write(dxf_file.getvalue())
                    tmp_path = tmp.name
                    st.session_state.tmp_dxf_path = tmp_path
                
                try:
                    reader = DXFReader(tmp_path)
                    reader.set_log_callback(lambda msg, lvl: None)
                    layers = reader.obtener_layers()
                    st.session_state.dxf_layers = layers
                    st.session_state.dxf_loaded = True
                    st.session_state.dxf_reader = reader
                    
                    st.info(f"📋 {len(layers)} layers encontrados")
                    
                    selected = st.multiselect(
                        "Seleccionar layers:",
                        options=layers,
                        default=layers[:3] if len(layers) >= 3 else layers,
                        key="layer_selector_main"
                    )
                    
                    if selected:
                        st.session_state.selected_layers = selected
                        if st.button("🚀 Construir Red", type="primary", use_container_width=True):
                            with st.spinner("Construyendo red..."):
                                lineas_raw = reader.extraer_lineas(selected)
                                if lineas_raw:
                                    factor = st.session_state.factor_conversion
                                    lineas = normalizar_coordenadas(lineas_raw, factor_conversion=factor)
                                    nodos_dict, tuberias = construir_red(lineas)
                                    
                                    red = RedHidraulica()
                                    for (x, y, z), nid in nodos_dict.items():
                                        red.agregar_nodo(Nodo(id=nid, x=x, y=y, z=z))
                                    for t in tuberias:
                                        red.agregar_tuberia(t)
                                    
                                    st.session_state.red = red
                                    
                                    if len(red.nodos) > 0:
                                        primer_nodo = list(red.nodos.keys())[0]
                                        red.nodo_entrada_id = primer_nodo
                                        red.nodos[primer_nodo].es_entrada = True
                                        ajustar_cotas_relativas(red)
                                    
                                    st.session_state.paso_actual = 2
                                    st.success(f"✅ Red: {len(red.nodos)} nodos, {len(red.tuberias)} tuberías")
                                    st.rerun()
                                else:
                                    st.error("No se encontraron líneas")
                except Exception as e:
                    st.error(f"Error: {e}")
                    st.session_state.dxf_loaded = False
                
                if st.session_state.tmp_dxf_path and os.path.exists(st.session_state.tmp_dxf_path):
                    try:
                        os.unlink(st.session_state.tmp_dxf_path)
                        st.session_state.tmp_dxf_path = None
                    except:
                        pass
        
        # ===== PASO 2: CONFIGURAR =====
        if st.session_state.red is not None:
            with st.expander("🔧 Configurar Nodos", expanded=(st.session_state.paso_actual == 2)):
                # Configuración 3D
                if st.button("🎯 Configurar en 3D", use_container_width=True):
                    nodos_data = [{"id": n.id, "x": n.x, "y": n.y, "z": n.z, 
                                   "es_entrada": n.es_entrada, "tipo_aparato": n.tipo_aparato, 
                                   "valvula_tipo": n.valvula_tipo, "valvula_cerrada": n.valvula_cerrada} 
                                  for n in st.session_state.red.nodos.values()]
                    
                    tuberias_data = []
                    for t in st.session_state.red.tuberias.values():
                        n1, n2 = st.session_state.red.nodos[t.nodo_inicio], st.session_state.red.nodos[t.nodo_fin]
                        tuberias_data.append({"id": t.id, "x1": n1.x, "y1": n1.y, "z1": n1.z, 
                                              "x2": n2.x, "y2": n2.y, "z2": n2.z})
                    
                    html_content = generar_html_config(nodos_data, tuberias_data)
                    st.components.v1.html(html_content, height=500, scrolling=True)
                    st.info("💡 Configure y descargue el archivo JSON")
                
                # Asignación manual
                nodo_ids = list(st.session_state.red.nodos.keys())
                entrada_actual = st.session_state.red.nodo_entrada_id
                nodo_entrada = st.selectbox(
                    "Nodo de entrada:",
                    options=nodo_ids,
                    index=nodo_ids.index(entrada_actual) if entrada_actual in nodo_ids else 0,
                    format_func=lambda x: f"Nodo {x}",
                    key="nodo_entrada_main"
                )
                
                if nodo_entrada != entrada_actual:
                    if st.button("✅ Establecer Entrada", use_container_width=True):
                        for n in st.session_state.red.nodos.values():
                            n.es_entrada = False
                        st.session_state.red.nodo_entrada_id = nodo_entrada
                        st.session_state.red.nodos[nodo_entrada].es_entrada = True
                        ajustar_cotas_relativas(st.session_state.red)
                        st.session_state.paso_actual = 3
                        st.success(f"✅ Nodo {nodo_entrada} como entrada")
                        st.rerun()
        
        # ===== PASO 3: ANALIZAR =====
        if st.session_state.red is not None and st.session_state.red.nodo_entrada_id is not None:
            with st.expander("⚙️ Parámetros y Análisis", expanded=(st.session_state.paso_actual == 3)):
                # Parámetros en dos columnas
                col1, col2 = st.columns(2)
                with col1:
                    presion = st.number_input(
                        "Presión (mca)",
                        value=st.session_state.presion_entrada,
                        min_value=1.0,
                        max_value=50.0,
                        step=0.5,
                        key="presion_main"
                    )
                    st.session_state.presion_entrada = presion
                
                with col2:
                    tipo_ocup = st.selectbox(
                        "Tipo edificación:",
                        options=list(TIPOS_OCUPACION_AGUA.keys()),
                        index=list(TIPOS_OCUPACION_AGUA.keys()).index(st.session_state.tipo_ocupacion) 
                             if st.session_state.tipo_ocupacion in TIPOS_OCUPACION_AGUA else 0,
                        key="tipo_ocup_main"
                    )
                    st.session_state.tipo_ocupacion = tipo_ocup
                
                # Velocidades
                col1, col2 = st.columns(2)
                with col1:
                    vel_min = st.number_input(
                        "Vel. mín (m/s)", 
                        value=st.session_state.vel_min, 
                        min_value=0.1, 
                        max_value=5.0, 
                        step=0.1,
                        key="vel_min_main"
                    )
                    st.session_state.vel_min = vel_min
                with col2:
                    vel_max = st.number_input(
                        "Vel. máx (m/s)", 
                        value=st.session_state.vel_max, 
                        min_value=0.5, 
                        max_value=10.0, 
                        step=0.1,
                        key="vel_max_main"
                    )
                    st.session_state.vel_max = vel_max
                
                # Diámetro máximo
                restringir = st.checkbox("Restringir diámetro máximo", key="restringir_main")
                if restringir:
                    diam_max = st.selectbox(
                        "Diámetro máximo:",
                        options=list(DIAMETROS_PAVCO.keys()),
                        key="diam_max_main"
                    )
                    st.session_state.diametro_maximo = DIAMETROS_PAVCO[diam_max]
                else:
                    st.session_state.diametro_maximo = None
                
                st.divider()
                
                # Botón ejecutar
                if st.button("🚀 EJECUTAR ANÁLISIS", type="primary", use_container_width=True):
                    with st.spinner("Ejecutando análisis..."):
                        analyzer = HydraulicAnalyzer(
                            st.session_state.red,
                            None,
                            st.session_state.diametro_maximo
                        )
                        analyzer.ejecutar()
                        st.session_state.analyzer = analyzer
                        
                        presiones = [n.presion_mca for n in st.session_state.red.nodos.values() if n.presion_mca is not None]
                        ug_acumulada = st.session_state.red.calcular_ug_acumulada()
                        ug_total = ug_acumulada.get(st.session_state.red.nodo_entrada_id, 0)
                        caudal_total = caudal_por_ug(ug_total, st.session_state.tipo_ocupacion)
                        
                        st.session_state.resultados = {
                            'presiones': presiones,
                            'ug_total': ug_total,
                            'caudal_total': caudal_total,
                            'cumple': min(presiones) >= PRESION_MIN_NORMA if presiones else False
                        }
                        
                        st.success("✅ Análisis completado")
                        st.rerun()
    
    # ============================================================
    # ÁREA PRINCIPAL - CONTENIDO DINÁMICO
    # ============================================================
    if st.session_state.red is None:
        # Pantalla de bienvenida
        st.title("💧 Hydro Domus Py")
        st.markdown("### Análisis Hidráulico para Redes Internas de Agua Potable")
        st.divider()
        
        col1, col2 = st.columns([2, 1])
        with col1:
            st.markdown("""
            **Pasos para realizar un análisis:**
            
            1. 📁 **Cargar archivo DXF** del plano hidrosanitario
            2. 📐 **Seleccionar las unidades** del dibujo (mm, cm, m, in, ft)
            3. 📐 **Seleccionar los layers** que contienen las tuberías
            4. 🔧 **Configurar nodos** (entrada, aparatos, válvulas)
            5. ⚙️ **Ajustar parámetros** de cálculo
            6. 🚀 **Ejecutar análisis** y revisar resultados
            """)
        
        with col2:
            st.info("""
            **📐 Características:**
            - Lectura DXF (LINE, LWPOLYLINE, POLYLINE, 3DPOLYLINE)
            - Análisis Darcy-Weisbach
            - Método de Hunter
            - Visualización 3D interactiva
            - Exportación a Excel
            """)
    
    else:
        # Mostrar resultados
        red = st.session_state.red
        resultados = st.session_state.resultados
        
        # ===== PESTAÑAS =====
        tab1, tab2, tab3, tab4, tab5 = st.tabs(["📊 Resumen", "🌐 Visualización 3D", "📈 Perfil", "📋 Tablas", "📦 Materiales"])
        
        with tab1:
            st.subheader("📊 Resumen del Análisis")
            
            if resultados:
                # Métricas en fila
                col1, col2, col3, col4, col5 = st.columns(5)
                col1.metric("🔢 Nodos", len(red.nodos))
                col2.metric("🔗 Tuberías", len(red.tuberias))
                col3.metric("🔧 Accesorios", len(red.accesorios))
                col4.metric("💧 Caudal total", f"{resultados['caudal_total']:.2f} L/s")
                col5.metric("📊 UG totales", f"{resultados['ug_total']:.0f}")
                
                st.divider()
                
                if resultados['presiones']:
                    p_min, p_max = min(resultados['presiones']), max(resultados['presiones'])
                    p_prom = sum(resultados['presiones']) / len(resultados['presiones'])
                    
                    col1, col2, col3, col4 = st.columns(4)
                    col1.metric("💪 Presión entrada", f"{st.session_state.presion_entrada:.1f} mca")
                    col2.metric("⬆️ Presión máxima", f"{p_max:.2f} mca")
                    col3.metric("⬇️ Presión mínima", f"{p_min:.2f} mca")
                    col4.metric("📊 Presión promedio", f"{p_prom:.2f} mca")
                    
                    if resultados['cumple']:
                        st.success(f"✅ CUMPLE NORMA NTC 1500 (Presión mínima ≥ {PRESION_MIN_NORMA} mca)")
                    else:
                        st.error(f"❌ NO CUMPLE NORMA NTC 1500 (Presión mínima: {p_min:.2f} mca < {PRESION_MIN_NORMA} mca)")
                
                st.divider()
                
                # Botones de exportación
                col1, col2 = st.columns(2)
                with col1:
                    if st.button("📊 Exportar a Excel", use_container_width=True):
                        try:
                            excel_data = exportar_excel(red)
                            b64 = base64.b64encode(excel_data).decode()
                            href = f'<a href="data:application/vnd.openxmlformats-officedocument.spreadsheetml.sheet;base64,{b64}" download="HydroDomusPy_Reporte_{datetime.now().strftime("%Y%m%d")}.xlsx">📥 Descargar Excel</a>'
                            st.markdown(href, unsafe_allow_html=True)
                            st.success("✅ Excel generado")
                        except Exception as e:
                            st.error(f"Error: {e}")
                
                with col2:
                    if st.button("💾 Guardar Configuración", use_container_width=True):
                        config = {
                            "nodo_entrada": red.nodo_entrada_id,
                            "nodos": [
                                {"id": n.id, "tipo_aparato": n.tipo_aparato, 
                                 "valvula_tipo": n.valvula_tipo, "valvula_cerrada": n.valvula_cerrada}
                                for n in red.nodos.values()
                            ],
                            "tipo_ocupacion": st.session_state.tipo_ocupacion,
                            "presion_entrada": st.session_state.presion_entrada
                        }
                        config_json = json.dumps(config, indent=2, ensure_ascii=False)
                        b64 = base64.b64encode(config_json.encode()).decode()
                        href = f'<a href="data:application/json;base64,{b64}" download="HydroDomusPy_config_{datetime.now().strftime("%Y%m%d")}.json">📥 Descargar Configuración</a>'
                        st.markdown(href, unsafe_allow_html=True)
                        st.success("✅ Configuración lista")
            else:
                st.info("Ejecute el análisis para ver los resultados")
        
        with tab2:
            st.subheader("🌐 Visualización 3D")
            if st.session_state.analyzer:
                fig = generate_3d_plot(red, st.session_state.presion_entrada)
                st.plotly_chart(fig, use_container_width=True, height=700)
            else:
                st.info("Ejecute el análisis para ver la visualización 3D")
        
        with tab3:
            st.subheader("📈 Perfil de Presiones")
            if st.session_state.analyzer and red.nodo_entrada_id is not None:
                fig = generar_perfil_presiones(red)
                if fig:
                    st.plotly_chart(fig, use_container_width=True)
                else:
                    st.warning("No se pudo generar el perfil")
            else:
                st.info("Ejecute el análisis para ver el perfil")
        
        with tab4:
            st.subheader("📋 Datos de la Red")
            tabla = st.selectbox("Seleccionar tabla:", ["Nodos", "Tuberías", "Accesorios"])
            
            if tabla == "Nodos":
                data = [{
                    "ID": n.id, "X": round(n.x, 2), "Y": round(n.y, 2), "Z": round(n.z, 2),
                    "Presión": round(n.presion_mca, 2) if n.presion_mca else None,
                    "Aparato": n.tipo_aparato or "-",
                    "Válvula": f"{n.valvula_tipo} ({'CERRADA' if n.valvula_cerrada else 'ABIERTA'})" if n.valvula_tipo else "-",
                    "Entrada": "✓" if n.es_entrada else ""
                } for n in red.nodos.values()]
                st.dataframe(data, use_container_width=True, height=400)
            
            elif tabla == "Tuberías":
                data = [{
                    "ID": t.id, "Desde": t.nodo_inicio, "Hasta": t.nodo_fin,
                    "Longitud": round(t.longitud_m, 2), "Diámetro": t.diametro_nominal_pulg,
                    "DI (mm)": round(t.diametro_mm, 2), "Caudal": round(t.caudal_lps, 3),
                    "Velocidad": round(t.velocidad_ms, 2), "Pérdida": round(t.perdida_mca, 3)
                } for t in red.tuberias.values()]
                st.dataframe(data, use_container_width=True, height=400)
            
            else:
                data = []
                for acc in red.accesorios:
                    diam = "N/A"
                    for t in red.tuberias.values():
                        if t.nodo_inicio == acc.nodo_id or t.nodo_fin == acc.nodo_id:
                            diam = t.diametro_nominal_pulg
                            break
                    data.append({
                        "ID": acc.id, "Tipo": acc.tipo.replace("_", " "),
                        "Nodo": acc.nodo_id, "DN": diam,
                        "Leq": round(acc.longitud_equivalente_m, 2),
                        "Pérdida": round(acc.perdida_mca, 4)
                    })
                st.dataframe(data, use_container_width=True, height=400)
        
        with tab5:
            st.subheader("📦 Estimación de Materiales")
            if st.session_state.analyzer:
                df_tuberias, df_accesorios, total_long, total_tramos, total_acc = generar_reporte_materiales(red)
                
                st.markdown("### 📏 Tuberías PVC")
                st.dataframe(df_tuberias, use_container_width=True)
                
                col1, col2 = st.columns(2)
                with col1:
                    st.metric("📐 Longitud total", f"{total_long:.2f} m")
                with col2:
                    st.metric("🔗 Tramos de 6m", f"{total_tramos}")
                
                st.divider()
                st.markdown("### 🔧 Accesorios")
                st.dataframe(df_accesorios, use_container_width=True)
                st.metric("📊 Total accesorios", f"{total_acc}")
            else:
                st.info("Ejecute el análisis para ver la estimación de materiales")

# ================================================================================
# EJECUTAR APLICACIÓN
# ================================================================================
if __name__ == "__main__":
    main()

# hydro_utils.py
# ================================================================================
#                            H Y D R O   U T I L S
# ================================================================================
#                    Utilidades para configuración y exportación
# ================================================================================

import streamlit as st
import json
import io
import math
import datetime
import networkx as nx
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from hydro_core import (
    UNIDADES_GASTO, DIAMETROS_PAVCO, diametro_a_numero,
    PRESION_MIN_NORMA
)

# ================================================================================
# CONFIGURACIÓN
# ================================================================================

def cargar_y_aplicar_configuracion(red, config_data):
    """
    Carga una configuración JSON y actualiza la red y la interfaz.
    Retorna True si fue exitoso, False en caso contrario.
    """
    try:
        # Resetear configuración actual
        for n in red.nodos.values():
            n.es_entrada = False
            n.tipo_aparato = ""
            n.valvula_tipo = ""
            n.valvula_cerrada = False
            n.valvula_apertura = 100.0
        
        # Asignar nodo de entrada
        nodo_entrada = config_data.get("nodo_entrada")
        if nodo_entrada in red.nodos:
            red.nodo_entrada_id = nodo_entrada
            red.nodos[nodo_entrada].es_entrada = True
        
        # Asignar propiedades de nodos
        for nodo_cfg in config_data.get("nodos", []):
            nid = nodo_cfg.get("id")
            if nid in red.nodos:
                if nodo_cfg.get("tipo_aparato"):
                    red.nodos[nid].tipo_aparato = nodo_cfg["tipo_aparato"]
                if nodo_cfg.get("valvula_tipo"):
                    red.nodos[nid].valvula_tipo = nodo_cfg["valvula_tipo"]
                    red.nodos[nid].valvula_apertura = nodo_cfg.get("valvula_apertura", 100.0)
                    red.nodos[nid].valvula_cerrada = nodo_cfg.get("valvula_cerrada", False)
                if nodo_cfg.get("es_entrada"):
                    red.nodos[nid].es_entrada = True
                    red.nodo_entrada_id = nid
        
        # Ajustar cotas relativas
        from hydro_core import ajustar_cotas_relativas
        ajustar_cotas_relativas(red)
        
        return True
    except Exception as e:
        st.error(f"Error cargando configuración: {e}")
        return False

def generar_configuracion_json(red, tipo_ocupacion, presion_entrada, unidad_dibujo):
    """Genera un objeto JSON con la configuración actual de la red"""
    config = {
        "nodo_entrada": red.nodo_entrada_id,
        "nodos": [
            {
                "id": n.id,
                "tipo_aparato": n.tipo_aparato,
                "valvula_tipo": n.valvula_tipo,
                "valvula_apertura": n.valvula_apertura,
                "valvula_cerrada": n.valvula_cerrada,
                "es_entrada": n.es_entrada
            }
            for n in red.nodos.values()
        ],
        "tipo_ocupacion": tipo_ocupacion,
        "presion_entrada": presion_entrada,
        "unidad_dibujo": unidad_dibujo,
        "fecha": datetime.datetime.now().isoformat()
    }
    return config

# ================================================================================
# EXPORTACIÓN A EXCEL
# ================================================================================

def generar_excel_bytes(red, ocupacion, presion_entrada, presion_minima):
    """
    Genera un archivo Excel con múltiples hojas:
    1. Metodología
    2. Resumen Ejecutivo
    3. Nodos
    4. Tuberías
    5. Accesorios y Materiales
    6. Perfil Crítico
    """
    wb = Workbook()
    wb.remove(wb.active)  # Quitar la hoja por defecto
    
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="FF2C3E50", end_color="FF2C3E50", fill_type="solid")
    title_font = Font(bold=True, size=14, color="FF1A5276")
    subtitle_font = Font(bold=True, size=12, color="FF2980B9")
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                         top=Side(style='thin'), bottom=Side(style='thin'))
    section_fill = PatternFill(start_color="FFD5DBDB", end_color="FFD5DBDB", fill_type="solid")
    
    # --- HOJA 1: METODOLOGÍA ---
    ws_metodo = wb.create_sheet("1. Metodología", 0)
    ws_metodo.merge_cells('A1:E1')
    ws_metodo.cell(row=1, column=1, value="HYDRO DOMUS PY - BASES DE CÁLCULO Y METODOLOGÍA").font = Font(bold=True, size=16, color="FF1A5276")
    
    textos_metodo = [
        ("1. Pérdidas por Fricción (Darcy-Weisbach)", 
         "El simulador utiliza la ecuación universal de Darcy-Weisbach para el cálculo exacto de las pérdidas de carga lineales. Esta ecuación es mecánicamente rigurosa y considera la conservación de energía (carga de velocidad) y la geometría (L/D)."),
        ("2. Factor de Fricción (Colebrook-White)", 
         "Para resolver la fricción (f), se emplea la ecuación implícita de Colebrook-White evaluada dinámicamente mediante métodos numéricos iterativos (fsolve). Considera la rugosidad absoluta del PVC y la viscosidad cinemática del agua calculando el Número de Reynolds en tiempo real."),
        ("3. Caudales de Diseño (Método de Hunter)", 
         "La asignación de caudales probabilísticos en la red se basa en el recuento topológico (Búsqueda en Anchura BFS) de Unidades de Gasto (UG) acumuladas, aplicando las curvas de Hunter adaptadas al tipo de edificación."),
        ("4. Pérdidas Localizadas (Long. Equivalentes)", 
         "Las pérdidas de energía en accesorios (codos, tees, válvulas) se computan transformando su resistencia local en una longitud de tubería equivalente (Leq), sumándola a la longitud real para el cálculo integral en Darcy-Weisbach."),
        ("5. Estrangulamiento de Válvulas", 
         "Se modela el cierre parcial de las válvulas afectando su longitud equivalente base mediante una penalización matemática (cuadrática), simulando la fuerte restricción de flujo al reducir el área de paso transversal.")
    ]
    
    row = 3
    for titulo, desc in textos_metodo:
        ws_metodo.cell(row=row, column=1, value=titulo).font = subtitle_font
        row += 1
        c = ws_metodo.cell(row=row, column=1, value=desc)
        c.alignment = Alignment(wrap_text=True)
        ws_metodo.merge_cells(f'A{row}:E{row}')
        ws_metodo.row_dimensions[row].height = 40
        row += 2
    ws_metodo.column_dimensions['A'].width = 30
    
    # --- HOJA 2: RESUMEN EJECUTIVO ---
    ws_resumen = wb.create_sheet("2. Resumen", 1)
    ws_resumen.merge_cells('A1:D1')
    ws_resumen.cell(row=1, column=1, value="RESUMEN EJECUTIVO DEL DISEÑO").font = Font(bold=True, size=16, color="FF1A5276")
    
    presiones = [n.presion_mca for n in red.nodos.values() if n.presion_mca is not None]
    velocidades = [t.velocidad_ms for t in red.tuberias.values()]
    ug_total = sum([UNIDADES_GASTO[n.tipo_aparato]["ug"] for n in red.nodos.values() if n.tipo_aparato])
    long_total = sum([t.longitud_m for t in red.tuberias.values()])
    
    row = 3
    ws_resumen.merge_cells(f'A{row}:D{row}')
    ws_resumen.cell(row=row, column=1, value="📐 ESTADÍSTICAS GLOBALES DEL SISTEMA").font = title_font
    ws_resumen.cell(row=row, column=1).fill = section_fill
    row += 1
    
    stats = [
        ["Tipo de Edificación", ocupacion],
        ["Total de Nodos", len(red.nodos)],
        ["Total de Tramos (Tuberías)", len(red.tuberias)],
        ["Total Aparatos Sanitarios", sum(1 for n in red.nodos.values() if n.tipo_aparato)],
        ["Total Unidades de Gasto (UG)", f"{ug_total} UG"],
        ["Longitud Total de Tubería", f"{long_total:.2f} m"],
        ["Total de Accesorios", len(red.accesorios)],
        ["Presión Inicial en Acometida", f"{presion_entrada:.2f} mca"],
        ["Presión Mínima Crítica", f"{min(presiones):.2f} mca" if presiones else "N/A"],
        ["Presión Máxima Registrada", f"{max(presiones):.2f} mca" if presiones else "N/A"],
        ["Velocidad Máxima Registrada", f"{max(velocidades):.2f} m/s" if velocidades else "N/A"],
        ["Velocidad Mínima Registrada", f"{min(velocidades):.2f} m/s" if velocidades else "N/A"]
    ]
    
    for k, v in stats:
        ws_resumen.cell(row=row, column=1, value=k).font = Font(bold=True)
        ws_resumen.cell(row=row, column=2, value=v)
        ws_resumen.cell(row=row, column=1).border = thin_border
        ws_resumen.cell(row=row, column=2).border = thin_border
        row += 1
    ws_resumen.column_dimensions['A'].width = 35
    ws_resumen.column_dimensions['B'].width = 25
    
    # --- HOJA 3: NODOS ---
    ws_nodos = wb.create_sheet("3. Nodos", 2)
    headers_n = ["ID Nodo", "X (m)", "Y (m)", "Cota Z (m)", "Presión (mca)", "Aparato", "Válvula", "Apertura (%)"]
    for col, h in enumerate(headers_n, 1):
        c = ws_nodos.cell(row=1, column=col, value=h)
        c.font, c.fill, c.border = header_font, header_fill, thin_border
        
    for r, n in enumerate(red.nodos.values(), 2):
        ws_nodos.cell(row=r, column=1, value=n.id).border = thin_border
        ws_nodos.cell(row=r, column=2, value=round(n.x, 2)).border = thin_border
        ws_nodos.cell(row=r, column=3, value=round(n.y, 2)).border = thin_border
        ws_nodos.cell(row=r, column=4, value=round(n.z, 2)).border = thin_border
        ws_nodos.cell(row=r, column=5, value=round(n.presion_mca, 2) if n.presion_mca else "-").border = thin_border
        ws_nodos.cell(row=r, column=6, value=n.tipo_aparato or "-").border = thin_border
        ws_nodos.cell(row=r, column=7, value=n.valvula_tipo or "-").border = thin_border
        ws_nodos.cell(row=r, column=8, value=n.valvula_apertura if n.valvula_tipo else "-").border = thin_border

    # --- HOJA 4: TUBERÍAS ---
    ws_tub = wb.create_sheet("4. Tuberías", 3)
    headers_t = ["ID Tubo", "Nodo Inicio", "Nodo Fin", "Longitud Real (m)", "Diámetro Nom.", "DI (mm)", "Caudal (L/s)", "Velocidad (m/s)", "Pérdida (mca)"]
    for col, h in enumerate(headers_t, 1):
        c = ws_tub.cell(row=1, column=col, value=h)
        c.font, c.fill, c.border = header_font, header_fill, thin_border
        
    for r, t in enumerate(red.tuberias.values(), 2):
        ws_tub.cell(row=r, column=1, value=t.id).border = thin_border
        ws_tub.cell(row=r, column=2, value=t.nodo_inicio).border = thin_border
        ws_tub.cell(row=r, column=3, value=t.nodo_fin).border = thin_border
        ws_tub.cell(row=r, column=4, value=round(t.longitud_m, 2)).border = thin_border
        ws_tub.cell(row=r, column=5, value=t.diametro_nominal_pulg).border = thin_border
        ws_tub.cell(row=r, column=6, value=round(t.diametro_mm, 2)).border = thin_border
        ws_tub.cell(row=r, column=7, value=round(t.caudal_lps, 3)).border = thin_border
        ws_tub.cell(row=r, column=8, value=round(t.velocidad_ms, 2)).border = thin_border
        ws_tub.cell(row=r, column=9, value=round(t.perdida_mca, 3)).border = thin_border

    # --- HOJA 5: ACCESORIOS Y MATERIALES ---
    ws_mat = wb.create_sheet("5. Accesorios", 4)
    ws_mat.merge_cells('A1:E1')
    ws_mat.cell(row=1, column=1, value="LISTADO DE ACCESORIOS DETECTADOS").font = title_font
    ws_mat.cell(row=1, column=1).fill = section_fill
    headers_m1 = ["ID", "Tipo de Accesorio", "Nodo", "Leq Equivalente (m)", "Pérdida Localizada (mca)"]
    for col, h in enumerate(headers_m1, 1):
        c = ws_mat.cell(row=2, column=col, value=h)
        c.font, c.fill, c.border = header_font, header_fill, thin_border
    
    row = 3
    acc_counts = {}
    for a in red.accesorios:
        ws_mat.cell(row=row, column=1, value=a.id).border = thin_border
        ws_mat.cell(row=row, column=2, value=a.tipo.replace("_", " ")).border = thin_border
        ws_mat.cell(row=row, column=3, value=a.nodo_id).border = thin_border
        ws_mat.cell(row=row, column=4, value=round(a.longitud_equivalente_m, 2)).border = thin_border
        ws_mat.cell(row=row, column=5, value=round(a.perdida_mca, 4)).border = thin_border
        row += 1
        acc_counts[a.tipo] = acc_counts.get(a.tipo, 0) + 1
        
    row += 2
    ws_mat.merge_cells(f'A{row}:D{row}')
    ws_mat.cell(row=row, column=1, value="CANTIDADES DE OBRA: TUBERÍAS (Factor Seg: 10%)").font = title_font
    ws_mat.cell(row=row, column=1).fill = section_fill
    row += 1
    headers_m2 = ["Diámetro Nominal", "Longitud Total (m)", "Tramos de 6m Necesarios"]
    for col, h in enumerate(headers_m2, 1):
        c = ws_mat.cell(row=row, column=col, value=h)
        c.font, c.fill, c.border = header_font, header_fill, thin_border
    row += 1
    
    longitudes_diam = {}
    for t in red.tuberias.values():
        longitudes_diam[t.diametro_nominal_pulg] = longitudes_diam.get(t.diametro_nominal_pulg, 0) + t.longitud_m
        
    for diam in sorted(longitudes_diam.keys(), key=lambda x: diametro_a_numero(x)):
        l = longitudes_diam[diam]
        tramos = math.ceil((l/6.0)*1.10)
        ws_mat.cell(row=row, column=1, value=diam).border = thin_border
        ws_mat.cell(row=row, column=2, value=round(l, 2)).border = thin_border
        ws_mat.cell(row=row, column=3, value=tramos).border = thin_border
        row += 1
        
    row += 2
    ws_mat.merge_cells(f'A{row}:B{row}')
    ws_mat.cell(row=row, column=1, value="ESTIMACIÓN DE CONSUMIBLES").font = title_font
    ws_mat.cell(row=row, column=1).fill = section_fill
    row += 1
    uniones_tubos = sum([max(0, math.ceil(t.longitud_m / 6.0) - 1) for t in red.tuberias.values()])
    uniones_acc = sum([3 if 'Tee' in k else 2 for k, v in acc_counts.items() for _ in range(v)])
    uniones_totales = int((uniones_tubos + uniones_acc) * 1.10)
    
    gal_pegamento = max(0.25, uniones_totales / 150)
    gal_limpiador = max(0.25, uniones_totales / 250)
    
    stats_cons = [
        ["Uniones Estimadas (+10%)", uniones_totales],
        ["Pegamento PVC (Galones)", round(gal_pegamento, 2)],
        ["Limpiador/Primer (Galones)", round(gal_limpiador, 2)]
    ]
    for k, v in stats_cons:
        ws_mat.cell(row=row, column=1, value=k).font = Font(bold=True)
        ws_mat.cell(row=row, column=2, value=v).border = thin_border
        ws_mat.cell(row=row, column=1).border = thin_border
        row += 1
    
    ws_mat.column_dimensions['B'].width = 25
    ws_mat.column_dimensions['D'].width = 20

    # --- HOJA 6: PERFIL CRÍTICO ---
    ws_perfil = wb.create_sheet("6. Perfil Crítico", 5)
    ws_perfil.merge_cells('A1:F1')
    ws_perfil.cell(row=1, column=1, value="ANÁLISIS DE LA RUTA CRÍTICA (TRAMO MÁS DESFAVORABLE)").font = title_font
    ws_perfil.cell(row=1, column=1).fill = section_fill
    
    headers_p = ["Nodo", "Distancia Acum. (m)", "Cota Elevación (m)", "Presión Nodo (mca)", "Tramo Previo", "Pérdida Tramo (mca)"]
    for col, h in enumerate(headers_p, 1):
        c = ws_perfil.cell(row=2, column=col, value=h)
        c.font, c.fill, c.border = header_font, header_fill, thin_border
        
    try:
        dist_nx = nx.single_source_dijkstra_path_length(red.grafo, red.nodo_entrada_id)
        nodo_lejano = max(dist_nx, key=dist_nx.get)
        camino = nx.shortest_path(red.grafo, red.nodo_entrada_id, nodo_lejano)
    except:
        camino = list(red.nodos.keys())[:10] if red.nodos else []

    if camino:
        dist_acum = 0
        ws_perfil.cell(row=3, column=1, value=camino[0]).border = thin_border
        ws_perfil.cell(row=3, column=2, value=0).border = thin_border
        ws_perfil.cell(row=3, column=3, value=round(red.nodos[camino[0]].z, 2)).border = thin_border
        ws_perfil.cell(row=3, column=4, value=round(red.nodos[camino[0]].presion_mca or 0, 2)).border = thin_border
        ws_perfil.cell(row=3, column=5, value="-").border = thin_border
        ws_perfil.cell(row=3, column=6, value="-").border = thin_border
        
        row = 4
        for i in range(1, len(camino)):
            n_prev = camino[i-1]
            n_curr = camino[i]
            tid = red._find_tuberia(n_prev, n_curr)
            tubo = red.tuberias[tid] if tid else None
            
            if tubo:
                dist_acum += tubo.longitud_m
                perdida = tubo.perdida_mca
                t_str = str(tubo.id)
            else:
                perdida = 0
                t_str = "N/A"
                
            nodo_obj = red.nodos[n_curr]
            ws_perfil.cell(row=row, column=1, value=n_curr).border = thin_border
            ws_perfil.cell(row=row, column=2, value=round(dist_acum, 2)).border = thin_border
            ws_perfil.cell(row=row, column=3, value=round(nodo_obj.z, 2)).border = thin_border
            ws_perfil.cell(row=row, column=4, value=round(nodo_obj.presion_mca or 0, 2)).border = thin_border
            ws_perfil.cell(row=row, column=5, value=t_str).border = thin_border
            ws_perfil.cell(row=row, column=6, value=round(perdida, 4)).border = thin_border
            row += 1
            
    ws_perfil.column_dimensions['B'].width = 20
    ws_perfil.column_dimensions['C'].width = 20
    ws_perfil.column_dimensions['D'].width = 20

    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()
